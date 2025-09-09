import asyncio
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path
from threading import Event
from urllib.parse import urlparse

import numpy as np
import pandas as pd
import pytz
import tqdm
import yaml
from fastapi.responses import FileResponse, PlainTextResponse
from loguru import logger
from nicegui import app, ui
from openpyxl.styles import Alignment, Border, Side
from playwright._impl._errors import Error, TimeoutError
from playwright.async_api import async_playwright
from playwright_stealth import Stealth

from beta_baza import parse_bet_baza
from parsers.marathonbet import get_players_links
from parsers.marathonbet import parse as marathonbet_parse

logger.add('logs/main.log')


def get_saved_url():
    url = None
    saved_url = Path('saved_url.yaml')
    if saved_url.exists():
        with saved_url.open(encoding='utf-8') as f:
            data = yaml.safe_load(f)
            if data:
                url = data.get('URL')
    return url


def save_url(url):
    if url:
        saved_url = Path('saved_url.yaml')
        data = {'URL': url}
        with saved_url.open('w', encoding='utf-8') as f:
            yaml.dump(data, f)


class MarathonbetURL:
    def __init__(self):
        self._value = None

    @property
    def value(self):
        _value = None
        if self._value:
            _value = self._value
        else:
            _value = get_saved_url()
        if not _value.endswith('/'):
            if _value:
                _value += '/'
        self._value = _value
        return self._value

    @value.setter
    def value_new_url(self, new_url):
        if new_url:
            self._value = new_url
            save_url(new_url)

    def new_url(self, new_url):
        if new_url:
            self._value = new_url
            save_url(new_url)


marathonbet_url = MarathonbetURL()
is_running = Event()


async def _parse():
    result = None
    singleton_lock_file = Path('browser/SingletonLock')
    if singleton_lock_file.exists():
        singleton_lock_file.unlink()
    async with Stealth().use_async(async_playwright()) as p:
        browser = await p.chromium.launch_persistent_context(
            user_data_dir='browser/',
            channel='chrome',
            headless=False,
            args=[
                '--start-maximized',
                '--disable-blink-features=AutomationControlled'
            ],
            base_url=marathonbet_url.value,
        )
        logger.info(f'Открываем {marathonbet_url.value}')
        page = await browser.new_page()
        page.set_default_timeout(180000)
        await page.goto('/')
        await page.wait_for_load_state()
        await page.goto('su')
        await page.wait_for_load_state()
        logger.info('Ждем окончания проверки браузера')
        try:
            while 'Just' in await page.title():
                await asyncio.sleep(1)
        except Error as exc:
            logger.exception(exc.message)
            page = browser.pages[-1]
            return PlainTextResponse('Во время обработки произошла ошибка. Попробуйте позже.')
        else:
            if urlparse(page.url).path != '/su/':
                await page.goto('su')
                await page.wait_for_load_state()
            try:
                await page.wait_for_selector(
                    '//table[@class="coupon-row-item"]',
                    timeout=180000
                )
            except TimeoutError as exc:
                logger.exception(exc.message)
                return PlainTextResponse('Вышло время ожидания страницы. Попробуйте позже.')
            else:
                logger.info('Собираем информацию по футболу за 24 часа')
                await page.get_by_text('Футбол').first.click()
                await page.get_by_text('24 часа').first.click()
                need_scroll = True
                attempts = 5
                content = ''
                while need_scroll:
                    await page.mouse.wheel(0, 3000)
                    await page.wait_for_load_state()
                    await page.wait_for_timeout(2000)
                    if content == await page.content():
                        if attempts:
                            attempts -= 1
                        else:
                            need_scroll = False
                    else:
                        content = await page.content()
                df_data = []
                players_links = get_players_links(await page.content())
                logger.info(f'Количество ссылок: {len(players_links)}')
                for player_link in tqdm.tqdm(players_links):
                    df_data_dict = dict()
                    player_page = await browser.new_page()
                    await player_page.goto(player_link)
                    await player_page.wait_for_load_state()
                    await player_page.wait_for_selector(
                        '//div[@class="block-market-wrapper"]',
                        timeout=180000
                    )
                    df_data_dict = marathonbet_parse(
                        await player_page.content(),
                        marathonbet_url.value + player_link[1:]
                    )
                    await player_page.close()
                    df_data.append(
                        df_data_dict
                    )
                await browser.close()
                if df_data:
                    logger.info(f'Собрано данных: {len(df_data)}')
                    df = pd.DataFrame.from_records(df_data)
                    now_msk = datetime.now(tz=pytz.timezone('Europe/Moscow'))
                    df['Дата слепка, МСК'] = now_msk
                    columns = [
                        'Ссылка',
                        'Название',
                        'Дата',
                        'Дата слепка, МСК',
                        '1',
                        'Х',
                        '2',
                        '1Х',
                        '12',
                        'Х2',
                        'Ф1(0)',
                        'Ф2(0)',
                        'ТМ(2.5)',
                        'ТБ(2.5)',
                        'ИТМ1(1.5)',
                        'ИТБ1(1.5)',
                        'ИТМ2(1.5)',
                        'ИТБ2(1.5)',
                        'ОЗ Да',
                        'ОЗ Нет',
                        'Гол оба тайма Да',
                        'Гол оба тайма Нет',
                        '_1_1',
                        '_1_Х',
                        '_1_2',
                        '_1_1Х',
                        '_1_12',
                        '_1_Х2',
                        '_1_Ф1(0)',
                        '_1_Ф2(0)',
                        '_1_ТМ(1.5)',
                        '_1_ТБ(1.5)',
                        '_1_ИТМ1(0.5)',
                        '_1_ИТБ1(0.5)',
                        '_1_ИТМ2(0.5)',
                        '_1_ИТБ2(0.5)',
                        '_2_1',
                        '_2_Х',
                        '_2_2',
                        '_2_1Х',
                        '_2_12',
                        '_2_Х2',
                        '_2_Ф1(0)',
                        '_2_Ф2(0)',
                        '_2_ТМ(1.5)',
                        '_2_ТБ(1.5)',
                        '_2_ИТМ1(0.5)',
                        '_2_ИТБ1(0.5)',
                        '_2_ИТМ2(0.5)',
                        '_2_ИТБ2(0.5)',
                    ]
                    df = df.reindex(columns=columns)
                    value_columns_start = columns.index('1')
                    # решаем проблему округления числа 1.285 в 1.29, а не 1.28 путем прибавления 0.0001
                    df.iloc[:, value_columns_start:] = (
                        df.iloc[:, value_columns_start:].astype(np.float64) + pow(10, -4)
                    ).round(2)
                    df['Дата'] = df['Дата'].dt.tz_localize(None)
                    df['Дата слепка, МСК'] = df['Дата слепка, МСК'].dt.tz_localize(None)
                    older_data = Path('storage') / Path('older.json')
                    if older_data.exists():
                        older_df = pd.read_json(older_data, date_unit='s')
                        older_df['Дата'] = older_df['Дата'].astype('datetime64[s]')
                        older_df['Дата слепка, МСК'] = older_df['Дата слепка, МСК'].astype('datetime64[s]')
                    else:
                        older_df = pd.DataFrame(columns=columns)
                    path = f'files/marathonbet_{now_msk.isoformat()}.xlsx'
                    if older_df.empty:
                        full_df = df
                    else:
                        full_df = pd.concat((df, older_df))
                    full_df.to_excel('files/debug.xlsx', index=False, columns=columns)
                    full_df = full_df.sort_values(
                        [
                            'Дата',
                            'Название',
                        ],
                        ascending=[False, True]
                    )
                    full_df['Double'] = full_df['Название'].duplicated()
                    full_df = full_df.reset_index(drop=True)
                    data = np.array(full_df[full_df['Double']].index.values)
                    ddiff = np.diff(data)
                    subArrays = np.split(data, np.where(ddiff != 1)[0]+1)

                    with pd.ExcelWriter(path, datetime_format='%d.%m.%y %H:%M') as writer:
                        full_df.to_excel(writer, index=False, startrow=1, columns=columns)
                        workbook = writer.book

                        sheet = workbook.active
                        for i in range(1, sheet.max_column + 1):
                            sheet.cell(2, i).alignment = Alignment(text_rotation=90)

                        match_index_start = columns.index('1') + 1
                        first_time_index_start = columns.index('_1_1') + 1
                        second_time_index_start = columns.index('_2_1') + 1

                        match_index_end = first_time_index_start - 1
                        first_time_index_end = second_time_index_start - 1
                        second_time_index_end = len(columns)

                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                        sheet.merge_cells(
                            start_row=1,
                            end_row=1,
                            start_column=match_index_start,
                            end_column=match_index_end
                        )
                        sheet.cell(1, match_index_start).value = 'Матч'
                        sheet.cell(1, match_index_start).alignment = Alignment(horizontal='center')
                        sheet.cell(1, match_index_start).border = thin_border

                        sheet.merge_cells(
                            start_row=1,
                            end_row=1,
                            start_column=first_time_index_start,
                            end_column=first_time_index_end
                        )
                        sheet.cell(1, first_time_index_start).value = '1 тайм'
                        sheet.cell(1, first_time_index_start).alignment = Alignment(horizontal='center')
                        sheet.cell(1, first_time_index_start).border = thin_border

                        sheet.merge_cells(
                            start_row=1,
                            end_row=1,
                            start_column=second_time_index_start,
                            end_column=second_time_index_end
                        )
                        sheet.cell(1, second_time_index_start).value = '2 тайм'
                        sheet.cell(1, second_time_index_start).alignment = Alignment(horizontal='center')
                        sheet.cell(1, second_time_index_start).border = thin_border

                        for i in range(first_time_index_start, first_time_index_end + 2):
                            sheet.cell(2, i - 1).value = sheet.cell(2, i - 1).value.replace('_1_', '')

                        for i in range(second_time_index_start, second_time_index_end + 2):
                            sheet.cell(2, i - 1).value = sheet.cell(2, i - 1).value.replace('_2_', '')

                        for subArray in subArrays:
                            if subArray.size > 0:
                                sheet.row_dimensions.group(subArray[0] + 3, subArray[-1] + 3, hidden=True)

                        workbook.save(path)
                    full_df.to_json(older_data, date_unit='s', date_format='iso')
                    result = FileResponse(
                        path=path,
                        filename=path
                    )
                else:
                    return PlainTextResponse('Не собрали данных')
        return result


@asynccontextmanager
async def parse_manager():
    try:
        if not is_running.is_set():
            is_running.set()
            response = await _parse()
        else:
            response = PlainTextResponse('В настоящий момент уже парсится сайт.')
        yield response
    finally:
        if is_running.is_set():
            is_running.clear()


@app.get('/parse')
async def parse():
    async with parse_manager() as response:
        return response


if __name__ in {"__main__", "__mp_main__"}:
    ui.page_title('Parser bet')
    ui.input('Ссылка:', value=marathonbet_url.value, on_change=lambda elem: marathonbet_url.new_url(elem.value))
    ui.link('Получить excel', '/parse', new_tab=True)
    ui.link('Получить данные Бет-База', '/parse_bet_baza', new_tab=True)
    ui.run(
        show=False
    )
