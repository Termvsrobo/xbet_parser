from abc import ABC, abstractmethod
from datetime import datetime, timedelta
from pathlib import Path
from threading import Event
from time import time
from typing import Any, Dict, List, Optional, Sequence, Union

import numpy as np
import pandas as pd
import pytz
from fastapi.responses import FileResponse, PlainTextResponse
from loguru import logger
from openpyxl.styles import Alignment, Border, Side
from pandas import DataFrame
from playwright.async_api import async_playwright
from playwright_stealth import Stealth
from pymongo.database import Database
from pymongo.results import InsertManyResult

from config import settings
from utils import (_get_db_instance, _handle_exists_collection,
                   _split_in_chunks, _validate_chunksize, get_saved_url,
                   save_url)


class ParserBase(ABC):

    @abstractmethod
    def start(self):
        raise NotImplementedError()

    @abstractmethod
    def stop(self):
        raise NotImplementedError()

    @abstractmethod
    def parse(self):
        raise NotImplementedError()

    @abstractmethod
    def parser_log_filter(self, record):
        raise NotImplementedError()


class Parser(ParserBase):
    def __init__(self, is_running: Event):
        self._value = None
        self.radio_period = '24 часа'
        self._count_links = None
        self._count_processed_links = None
        self._elapsed_time = None
        self._eta = None
        self._status = None
        self._is_running = is_running

        self._url_config_file = f'{self.name}_url.yaml'
        self._url = None
        self._path = None
        self.logger = logger
        self.logger.add(f'logs/{self.name}.log', filter=self.parser_log_filter)

    def read_mongo(
        self,
        collection: str,
        query: List[Dict[str, Any]],
        db: Union[str, Database],
        index_col: Optional[Union[str, List[str]]] = None,
        extra: Optional[Dict[str, Any]] = None,
        chunksize: Optional[int] = None
    ) -> DataFrame:
        """
        Read MongoDB query into a DataFrame.

        Returns a DataFrame corresponding to the result set of the query.
        Optionally provide an `index_col` parameter to use one of the
        columns as the index, otherwise default integer index will be used.

        Parameters
        ----------
        collection : str
            Mongo collection to select for querying
        query : list
            Must be an aggregate query.
            The input will be passed to pymongo `.aggregate`
        db : pymongo.database.Database or database string URI
            The database to use
        index_col : str or list of str, optional, default: None
            Column(s) to set as index(MultiIndex).
        extra : dict, optional, default: None
            List of parameters to pass to aggregate method.
        chunksize : int, default None
            If specified, return an iterator where `chunksize` is the number of
            docs to include in each chunk.
        Returns
        -------
        Dataframe
        """
        params = {}
        if chunksize is not None:
            _validate_chunksize(chunksize)

            params['batchSize'] = chunksize
        db = _get_db_instance(db)
        if extra is None:
            extra = {}

        if extra.get('batchSize') is not None:
            if chunksize is not None:
                raise ValueError("Either chunksize or batchSize must be provided, not both")

        return DataFrame.from_records(
            db[collection].aggregate(query, **{**params, **extra}),
            index=index_col)

    def to_mongo(
        self,
        frame: DataFrame,
        name: str,
        db: Union[str, Database],
        if_exists: Optional[str] = "fail",
        index: Optional[bool] = True,
        index_label: Optional[Union[str, Sequence[str]]] = None,
        chunksize: Optional[int] = None,
    ) -> Union[List[InsertManyResult], InsertManyResult]:
        """
        Write records stored in a DataFrame to a MongoDB collection.

        Parameters
        ----------
        frame : DataFrame, Series
        name : str
            Name of collection.
        db : pymongo.database.Database or database string URI
            The database to write to
        if_exists : {'fail', 'replace', 'append'}, default 'fail'
            - fail: If table exists, do nothing.
            - replace: If table exists, drop it, recreate it, and insert data.
            - append: If table exists, insert data. Create if does not exist.
        index : boolean, default True
            Write DataFrame index as a column.
        index_label : str or sequence, optional
            Column label for index column(s). If None is given (default) and
            `index` is True, then the index names are used.
            A sequence should be given if the DataFrame uses MultiIndex.
        chunksize : int, optional
            Specify the number of rows in each batch to be written at a time.
            By default, all rows will be written at once.
        """
        db = _get_db_instance(db)
        _handle_exists_collection(name, if_exists, db)
        records = frame.to_dict('records')
        if index is True:
            idx = frame.index
            idx_name = idx.name
            idx_data = idx.tolist()
            for i, record in enumerate(records):
                if index_label is None and idx_name is not None:
                    record[idx_name] = idx_data[i]
        if chunksize is not None:
            _validate_chunksize(chunksize)
            result_insert_many = []
            for chunk in _split_in_chunks(records, chunksize):
                result_insert_many.append(db[name].insert_many(chunk))
            return result_insert_many
        return db[name].insert_many(records)

    @property
    def name(self):
        return self.__class__.__name__.lower()

    @property
    def path(self):
        if self._path:
            return self._path
        else:
            return None

    @path.setter
    def path(self, new_path):
        if self._path and self._path != new_path:
            self.delete_older_file()
        self._path = new_path

    def delete_older_file(self):
        if self._path:
            Path(self._path).unlink(missing_ok=True)
            self._path = None

    @property
    def url(self):
        _url = None
        if self._url:
            _url = self._url
        else:
            _url = get_saved_url(self._url_config_file)
        if _url and not _url.endswith('/'):
            if _url:
                _url += '/'
        self._url = _url
        return self._url

    @url.setter
    def url(self, new_url):
        if new_url:
            self._url = new_url
            save_url(self._url_config_file, new_url)

    @property
    def is_running(self):
        return self._is_running.is_set()

    def start(self):
        self.delete_older_file()
        self._elapsed_time = time()

    def stop(self):
        self._count_processed_links = None
        self._count_links = None
        self._elapsed_time = None
        self._eta = None
        self._status = None

    @property
    def status(self):
        if self._status:
            return f'Статус: {self._status}'
        else:
            return 'Статус: --'

    @status.setter
    def status(self, value: str):
        if value:
            self._status = value
        else:
            self._status = None

    @property
    def elapsed_time(self):
        if self._elapsed_time:
            result = []
            td = timedelta(seconds=time() - self._elapsed_time)
            hours, remainder = divmod(td.seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            if hours:
                result.append(f'{hours} ч.')
            if minutes:
                result.append(f'{minutes} мин.')
            if seconds:
                result.append(f'{seconds} сек.')
            return f'Прошло {", ".join(result)}'
        else:
            return 'Прошло -- сек.'

    @property
    def count_links(self):
        if self._count_links:
            return f'Количество ссылок: {str(self._count_links)}'
        else:
            if self.is_running:
                return 'Количество ссылок: Вычисляем'
            else:
                return 'Количество ссылок: -'

    @count_links.setter
    def count_links(self, value: int):
        if value:
            self._count_links = value
        else:
            self._count_links = None

    @property
    def count_processed_links(self):
        if self._count_processed_links and self._count_links:
            percent = round(self._count_processed_links / self._count_links * 100, 2)
            return f'Обработано ссылок: {str(self._count_processed_links)} ({percent} %)'
        else:
            if self.is_running:
                return 'Обработано ссылок: Вычисляем'
            else:
                return 'Обработано ссылок: -'

    @count_processed_links.setter
    def count_processed_links(self, value: int):
        if value:
            self._count_processed_links = value
        else:
            self._count_processed_links = None

    def tqdm(self, links):
        start = time()
        for i, link in enumerate(links):
            yield link
            self.count_processed_links = i
            end = time()
            delta = end - start
            self._eta = (len(links) - (i + 1)) * delta
            start = end

    @property
    def eta(self):
        if self._eta:
            result = []
            td = timedelta(seconds=self._eta)
            hours, remainder = divmod(td.seconds, 3600)
            minutes, seconds = divmod(remainder, 60)
            if hours:
                result.append(f'{hours} ч.')
            if minutes:
                result.append(f'{minutes} мин.')
            if seconds:
                result.append(f'{seconds} сек.')
            return f'Осталось примерно {", ".join(result)}'
        else:
            return 'Осталось -- сек.'

    def get_file_response(self, df_data):
        result = None
        if df_data:
            msg = f'Собрано данных: {len(df_data)}'
            logger.info(msg)
            self.status = msg
            df = pd.DataFrame.from_records(df_data)
            now_msk = datetime.now(tz=pytz.timezone('Europe/Moscow'))
            df['Дата слепка, МСК'] = now_msk
            columns = [
                'Ссылка',
                'Страна',
                'Лига',
                'Команда 1',
                'Команда 2',
                'Дата',
                'Дата слепка, МСК',
                '1',
                'Х',
                '2',
                '1Х',
                '12',
                'Х2',
                'Ф1(-1.5)',
                'Ф1(-1.0)',
                'Ф1(0)',
                'Ф1(+1.0)',
                'Ф1(+1.5)',
                'Ф2(-1.5)',
                'Ф2(-1.0)',
                'Ф2(0)',
                'Ф2(+1.0)',
                'Ф2(+1.5)',
                'ТМ(1.5)',
                'ТМ(2.0)',
                'ТМ(2.5)',
                'ТМ(3.0)',
                'ТМ(3.5)',
                'ТБ(1.5)',
                'ТБ(2.0)',
                'ТБ(2.5)',
                'ТБ(3.0)',
                'ТБ(3.5)',
                'ИТМ1(1.0)',
                'ИТМ1(1.5)',
                'ИТМ1(2.0)',
                'ИТБ1(1.0)',
                'ИТБ1(1.5)',
                'ИТБ1(2.0)',
                'ИТМ2(1.0)',
                'ИТМ2(1.5)',
                'ИТМ2(2.0)',
                'ИТБ2(1.0)',
                'ИТБ2(1.5)',
                'ИТБ2(2.0)',
                'ОЗ Да',
                'ОЗ Нет',
                'Гол оба тайма Да',
                'Гол оба тайма Нет',
                '_1_1',
                '_1_Х',
                '_1_2',
                '_1_1Х',
                '_1_12',
                '_1_Х2',
                '_1_Ф1(-1.0)',
                '_1_Ф1(0)',
                '_1_Ф1(+1.0)',
                '_1_Ф2(-1.0)',
                '_1_Ф2(0)',
                '_1_Ф2(+1.0)',
                '_1_ТМ(0.5)',
                '_1_ТМ(1.0)',
                '_1_ТМ(1.5)',
                '_1_ТМ(2.0)',
                '_1_ТМ(2.5)',
                '_1_ТБ(0.5)',
                '_1_ТБ(1.0)',
                '_1_ТБ(1.5)',
                '_1_ТБ(2.0)',
                '_1_ТБ(2.5)',
                '_1_ИТМ1(0.5)',
                '_1_ИТМ1(1.0)',
                '_1_ИТМ1(1.5)',
                '_1_ИТБ1(0.5)',
                '_1_ИТБ1(1.0)',
                '_1_ИТБ1(1.5)',
                '_1_ИТМ2(0.5)',
                '_1_ИТМ2(1.0)',
                '_1_ИТМ2(1.5)',
                '_1_ИТБ2(0.5)',
                '_1_ИТБ2(1.0)',
                '_1_ИТБ2(1.5)',
                '_2_1',
                '_2_Х',
                '_2_2',
                '_2_1Х',
                '_2_12',
                '_2_Х2',
                '_2_Ф1(-1.0)',
                '_2_Ф1(0)',
                '_2_Ф1(+1.0)',
                '_2_Ф2(-1.0)',
                '_2_Ф2(0)',
                '_2_Ф2(+1.0)',
                '_2_ТМ(0.5)',
                '_2_ТМ(1.0)',
                '_2_ТМ(1.5)',
                '_2_ТМ(2.0)',
                '_2_ТМ(2.5)',
                '_2_ТБ(0.5)',
                '_2_ТБ(1.0)',
                '_2_ТБ(1.5)',
                '_2_ТБ(2.0)',
                '_2_ТБ(2.5)',
                '_2_ИТМ1(0.5)',
                '_2_ИТМ1(1.0)',
                '_2_ИТМ1(1.5)',
                '_2_ИТБ1(0.5)',
                '_2_ИТБ1(1.0)',
                '_2_ИТБ1(1.5)',
                '_2_ИТМ2(0.5)',
                '_2_ИТМ2(1.0)',
                '_2_ИТМ2(1.5)',
                '_2_ИТБ2(0.5)',
                '_2_ИТБ2(1.0)',
                '_2_ИТБ2(1.5)'
            ]
            df = df.reindex(columns=columns)
            value_columns_start = columns.index('1')
            # решаем проблему округления числа 1.285 в 1.29, а не 1.28 путем прибавления 0.0001
            df.iloc[:, value_columns_start:] = (
                df.iloc[:, value_columns_start:].astype(np.float64) + pow(10, -4)
            ).round(2)
            df['Дата'] = df['Дата'].dt.tz_localize(None)
            df['Дата слепка, МСК'] = df['Дата слепка, МСК'].dt.tz_localize(None)
            older_df = pd.DataFrame(columns=columns)
            if not settings.DEBUG:
                older_df = self.read_mongo('History', [], settings.MONGO_URL.encoded_string())
            self.to_mongo(df, 'History', settings.MONGO_URL.encoded_string(), if_exists='append', index=False)
            self.path = f'files/{self.name}_{now_msk.isoformat()}.xlsx'
            if older_df.empty:
                full_df = df
            else:
                full_df = pd.concat((df, older_df))
            if settings.DEBUG:
                full_df.to_excel('files/debug.xlsx', index=False, columns=columns)
            full_df = full_df.sort_values(
                [
                    'Дата',
                    'Команда 1',
                    'Команда 2',
                ],
                ascending=[False, True, True]
            )
            full_df['Double'] = full_df[['Команда 1', 'Команда 2', 'Дата']].duplicated()
            full_df = full_df.reset_index(drop=True)
            data = np.array(full_df[full_df['Double']].index.values)
            ddiff = np.diff(data)
            subArrays = np.split(data, np.where(ddiff != 1)[0]+1)

            with pd.ExcelWriter(self.path, datetime_format='%d.%m.%y %H:%M') as writer:
                full_df.to_excel(writer, index=False, startrow=1, columns=columns)
                workbook = writer.book

                sheet = workbook.active
                for i in range(1, sheet.max_column + 1):
                    sheet.cell(2, i).alignment = Alignment(text_rotation=90)

                match_index_start = columns.index('1') + 1
                first_time_index_start = columns.index('_1_1') + 1
                second_time_index_start = columns.index('_2_1') + 1

                match_index_end = first_time_index_start - 1
                first_time_index_end = second_time_index_start - 1
                second_time_index_end = len(columns)

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                sheet.merge_cells(
                    start_row=1,
                    end_row=1,
                    start_column=match_index_start,
                    end_column=match_index_end
                )
                sheet.cell(1, match_index_start).value = 'Матч'
                sheet.cell(1, match_index_start).alignment = Alignment(horizontal='center')
                sheet.cell(1, match_index_start).border = thin_border

                sheet.merge_cells(
                    start_row=1,
                    end_row=1,
                    start_column=first_time_index_start,
                    end_column=first_time_index_end
                )
                sheet.cell(1, first_time_index_start).value = '1 тайм'
                sheet.cell(1, first_time_index_start).alignment = Alignment(horizontal='center')
                sheet.cell(1, first_time_index_start).border = thin_border

                sheet.merge_cells(
                    start_row=1,
                    end_row=1,
                    start_column=second_time_index_start,
                    end_column=second_time_index_end
                )
                sheet.cell(1, second_time_index_start).value = '2 тайм'
                sheet.cell(1, second_time_index_start).alignment = Alignment(horizontal='center')
                sheet.cell(1, second_time_index_start).border = thin_border

                for i in range(first_time_index_start, first_time_index_end + 2):
                    sheet.cell(2, i - 1).value = sheet.cell(2, i - 1).value.replace('_1_', '')

                for i in range(second_time_index_start, second_time_index_end + 2):
                    sheet.cell(2, i - 1).value = sheet.cell(2, i - 1).value.replace('_2_', '')

                for subArray in subArrays:
                    if subArray.size > 0:
                        sheet.row_dimensions.group(subArray[0] + 3, subArray[-1] + 3, hidden=True)

                workbook.save(self.path)
            result = FileResponse(
                self.path,
                filename=f'{self.name}_{now_msk.isoformat()}.xlsx'
            )
        else:
            result = PlainTextResponse('Не собрали данных')
        return result


class BrowserManager:
    def __init__(self, is_running: Event, parser: Parser):
        self._is_running = is_running
        self._parser = parser
        self._ctx_browser = None

    @property
    def parser(self):
        return self._parser

    @property
    def is_running(self):
        return self._is_running.is_set()

    async def __aenter__(self):
        if not self.is_running:
            singleton_lock_file = Path('browser/SingletonLock')
            if singleton_lock_file.exists():
                singleton_lock_file.unlink()
            self._ctx_browser = Stealth().use_async(async_playwright())
            p = await self._ctx_browser.__aenter__()
            browser = await p.chromium.launch_persistent_context(
                user_data_dir='browser/',
                channel='chrome',
                headless=False,
                args=[
                    '--start-maximized',
                    '--disable-blink-features=AutomationControlled'
                ],
                base_url=self.parser.url,
                screen={
                    "width": 1920,
                    "height": 1080
                },
                viewport={
                    "width": 1920,
                    "height": 1080
                }
            )
            self._is_running.set()
            return browser

    async def parse(self, browser):
        self.parser.start()
        result = await self.parser.parse(browser)
        return result

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        self._is_running.clear()
        self.parser.stop()
        await self._ctx_browser.__aexit__(exc_type, exc_val, exc_tb)
