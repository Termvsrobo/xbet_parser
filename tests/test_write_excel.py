from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment


def test_write_excel_header():
    df = pd.DataFrame(
        {
            'test 1': [0, 1.234234, 6.234],
            'test 2': [1, 1.234234, 6.234],
            'test 3': [2, 1.234234, 6.234],
        }
    )
    df = df.round(2)
    path = f'test_{datetime.now().isoformat()}.xlsx'
    with pd.ExcelWriter(path) as writer:
        df.to_excel(writer, index=False)
        workbook = writer.book

        sheet = workbook.active
        for i in range(1, sheet.max_column + 1):
            sheet.cell(1, i).alignment = Alignment(text_rotation=90)
        workbook.save(path)


def test_group_rows():
    df = pd.DataFrame(
        {
            'Name': ['n1', 'n2', 'n2', 'n3', 'n4', 'n2', 'n4', 'n2'],
            'Value': [1, 1.234234, 6.234, 4, 123.0, 12., 82., 54.],
            'Rate': [2, 1.234234, 6.234, 4, 123.0, 14., 38., 22.],
        }
    )
    columns = df.columns.tolist()
    path = f'test_{datetime.now().isoformat()}.xlsx'
    with pd.ExcelWriter(path) as writer:
        df = df.sort_values(['Name', 'Value'])
        df['Double'] = df['Name'].duplicated()
        df = df.reset_index(drop=True)

        data = np.array(df[df['Double'] is True].index.values)
        ddiff = np.diff(data)
        subArrays = np.split(data, np.where(ddiff != 1)[0]+1)

        df.to_excel(writer, index=False, columns=columns)
        workbook = writer.book

        sheet = workbook.active
        for i in range(1, sheet.max_column + 1):
            sheet.cell(1, i).alignment = Alignment(text_rotation=90)
        for subArray in subArrays:
            sheet.row_dimensions.group(subArray[0] + 2, subArray[-1] + 2, hidden=True)
        workbook.save(path)


def test_merge_header_cells():
    df = pd.DataFrame(
        {
            'Name': ['n1', 'n2', 'n2', 'n3', 'n4'],
            'Value': [1, 1.234234, 6.234, 4, 123.0],
            'Rate': [2, 1.234234, 6.234, 4, 123.0],
            'Value 2': [3, 1.234234, 6.234, 4, 123.0],
            'Rate 2': [4, 1.234234, 6.234, 4, 123.0],
        }
    )
    path = f'test_{datetime.now().isoformat()}.xlsx'
    with pd.ExcelWriter(path) as writer:
        df.to_excel(writer, index=False, startrow=1)
        workbook = writer.book

        sheet = workbook.active
        for i in range(1, sheet.max_column + 1):
            sheet.cell(2, i).alignment = Alignment(text_rotation=90)
        sheet.row_dimensions.group(5, 5)
        # sheet.merge_cells('B1:C1')
        # sheet.merge_cells('D1:E1')
        sheet.merge_cells(start_row=1, end_row=1, start_column=2, end_column=3)
        sheet.cell(1, 2).value = 'Test1'
        sheet.cell(1, 2).alignment = Alignment(horizontal='center')
        sheet.merge_cells(start_row=1, end_row=1, start_column=4, end_column=5)
        sheet.cell(1, 4).value = 'Test2'
        sheet.cell(1, 4).alignment = Alignment(horizontal='center')
        workbook.save(path)
