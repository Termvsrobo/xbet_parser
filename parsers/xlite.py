import asyncio
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse, urlunparse

import httpx
import numpy as np
import pandas as pd
import pytz
from bs4 import BeautifulSoup
from fastapi.responses import FileResponse, PlainTextResponse
from loguru import logger
from openpyxl.styles import Alignment, Border, Side

from base import Parser

logger.add('logs/xlite.log')

DEBUG = True


class XLiteParser(Parser):
    def get_players_links(self, page_content):
        soup = BeautifulSoup(page_content, 'html.parser')
        players_links = set()
        players = soup.find_all(
            lambda tag: tag.name == 'a'
            and tag.get('class') == ['ui-sports-event__link']
        )
        for player in players:
            players_links.add(player.attrs.get('href'))
        return players_links

    async def get_all_ids(self, min_offset: Optional[int] = None):
        result = []
        params = {
            'sports': 1,
            'country': 1,
            'virtualSports': True,
            'gr': 285,
            'groupChamps': True
        }
        if min_offset:
            params['minOffset'] = min_offset
        async with httpx.AsyncClient(
            headers={
                'User-Agent': 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Mobile Safari/537.36'  # noqa:E501
            }
        ) as client:
            scheme, domain, _, _, _, _ = urlparse(self.url)
            url = urlunparse((
                scheme,
                domain,
                '/service-api/LineFeed/GetSportsShortZip',
                None,
                None,
                None
            ))
            response = await client.get(url, params=params)
            data = response.json()
            if 'Value' in data:
                data_value = data['Value']
                football_data = next(filter(lambda x: x.get('N', '') == 'Футбол' and 'L' in x, data_value), None)
                list_champs = list(map(lambda x: x.get('LI'), filter(lambda x: 'SC' not in x, football_data['L'])))
                added_list_champs = [
                    sc.get('LI')
                    for i in filter(lambda x: 'SC' in x, football_data['L'])
                    for sc in i.get('SC', [])
                ]
                params = {
                    'sports': 1,
                    'champs': ','.join(map(str, sorted(list_champs))),
                    'country': 1,
                    'virtualSports': True,
                    'gr': 285,
                    'groupChamps': True
                }
                if min_offset:
                    params['minOffset'] = min_offset
                list_champs_response = await client.get(
                    url,
                    params=params
                )
                if list_champs_response.status_code == 200:
                    list_champs_data = list_champs_response.json()
                    if 'Value' in list_champs_data:
                        list_champs_data = list_champs_data['Value']
                        _list_champs_data = next(
                            filter(lambda x: x.get('N', '') == 'Футбол' and 'L' in x, list_champs_data),
                            None
                        )
                        for l_list in _list_champs_data.get('L', []):
                            for g_list in l_list.get('G', []):
                                ci = g_list.get('CI')
                                if ci:
                                    result.append(ci)
                for li in sorted(added_list_champs):
                    params = {
                        'sports': 1,
                        'champs': li,
                        'country': 1,
                    }
                    if min_offset:
                        params['minOffset'] = min_offset
                    champ_response = await client.get(
                        url,
                        params=params
                    )
                    if champ_response.status_code == 200:
                        champ_data = champ_response.json()
                        if 'Value' in champ_data:
                            champ_value = champ_data['Value']
                            _champ_data = next(
                                filter(lambda x: x.get('N', '') == 'Футбол' and 'L' in x, champ_value),
                                None
                            )
                            for l_list in _champ_data.get('L', []):
                                for g_list in l_list.get('G', []):
                                    ci = g_list.get('CI')
                                    if ci:
                                        result.append(ci)
        return sorted(result)

    async def _parse(self, page_id):
        result_dict = defaultdict(lambda: None)
        country_name = None
        league_name = None
        name_players = [None] * 2
        date_game = None
        keys = {
            1: {
                1: {
                    1: 'P1',
                    2: 'X',
                    3: 'P2',
                }
            },
            2: {
                3: {
                    7: ('F1_0', {'default': 0, 'times': 0}),
                    8: ('F2_0', {'default': 0, 'times': 0}),
                }
            },
            8: {
                2: {
                    4: '_1X',
                    5: '_12',
                    6: '_2X',
                }
            },
            15: {
                5: {
                    11: ('IT1_bol', {'default': 1.5, 'times': 0.5}),
                    12: ('IT1_men', {'default': 1.5, 'times': 0.5}),
                }
            },
            17: {
                4: {
                    9: ('TB', {'default': 2.5, 'times': 1.5}),
                    10: ('TM', {'default': 2.5, 'times': 1.5}),
                }
            },
            19: {
                21: {
                    180: 'ALL_win_yes',
                    181: 'ALL_win_no',
                }
            },
            32: {
                42: {
                    478: 'ALL_times_yes',
                    479: 'ALL_times_no',
                }
            },
            62: {
                6: {
                    13: ('IT2_bol', {'default': 1.5, 'times': 0.5}),
                    14: ('IT2_men', {'default': 1.5, 'times': 0.5}),
                }
            },
        }
        df_data_dict = dict()
        async with httpx.AsyncClient() as client:
            if page_id:
                scheme, domain, _, _, _, _ = urlparse(self.url)
                url = urlunparse((
                    scheme,
                    domain,
                    '/service-api/LineFeed/GetGameZip',
                    None,
                    None,
                    None
                ))
                response = await client.get(
                    url,
                    params={
                        'id': page_id,
                        'isSubGames': True,
                        'GroupEvents': True,
                        'countevents': 750,
                        'grMode': 4,
                        'topGroups': '',
                        'country': 1,
                        'marketType': 1,
                        'isNewBuilder': True
                    }
                )
                data = response.json()
                data_value = data['Value']
                if data_value:
                    league_header_data = data_value['L'].split()
                    if len(league_header_data) == 2:
                        country_name = league_header_data[0]
                        league_name = league_header_data[1]
                    elif len(league_header_data) > 2:
                        country_name = league_header_data[0]
                        league_name = ' '.join(league_header_data[1:])
                    name_players = [data_value['O1'], data_value['O2']]
                    date_game = datetime.fromtimestamp(data_value['S'])
                    for ge in data_value['GE']:
                        for e in ge['E']:
                            for row in e:
                                key = keys.get(row['G'], {}).get(row['GS'], {}).get(row['T'])
                                if isinstance(key, str):
                                    result_dict[key] = row['C']
                                elif isinstance(key, tuple):
                                    _k, _p = key
                                    _p = _p['default']
                                    if row.get('P', 0) == _p:
                                        result_dict[_k] = row['C']

                    first_time_page_id = next(
                        filter(lambda x: x['PN'] == '1-й тайм' and not x['TG'], data_value.get('SG', [])),
                        None
                    )
                    if first_time_page_id:
                        first_time_page_id = first_time_page_id['CI']
                        first_time_response = await client.get(
                            url,
                            params={
                                'id': first_time_page_id,
                                'isSubGames': True,
                                'GroupEvents': True,
                                'countevents': 1750,
                                'grMode': 4,
                                'topGroups': '',
                                'country': 1,
                                'marketType': 1,
                                'isNewBuilder': True
                            }
                        )
                        first_time_data = first_time_response.json()
                        first_time_data_value = first_time_data['Value']
                        if first_time_data_value:
                            for ge in first_time_data_value['GE']:
                                for e in ge['E']:
                                    for row in e:
                                        key = keys.get(row['G'], {}).get(row['GS'], {}).get(row['T'])
                                        if isinstance(key, str):
                                            result_dict['1_time_' + key] = row['C']
                                        elif isinstance(key, tuple):
                                            _k, _p = key
                                            _p = _p['times']
                                            if row.get('P', 0) == _p:
                                                result_dict['1_time_' + _k] = row['C']

                    second_time_page_id = next(
                        filter(lambda x: x['PN'] == '2-й тайм' and not x['TG'], data_value.get('SG', [])),
                        None
                    )
                    if second_time_page_id:
                        second_time_page_id = second_time_page_id['CI']
                        second_time_response = await client.get(
                            url,
                            params={
                                'id': second_time_page_id,
                                'isSubGames': True,
                                'GroupEvents': True,
                                'countevents': 750,
                                'grMode': 4,
                                'topGroups': '',
                                'country': 1,
                                'marketType': 1,
                                'isNewBuilder': True
                            }
                        )
                        second_time_data = second_time_response.json()
                        second_time_data_value = second_time_data['Value']
                        if second_time_data_value:
                            for ge in second_time_data_value['GE']:
                                for e in ge['E']:
                                    for row in e:
                                        key = keys.get(row['G'], {}).get(row['GS'], {}).get(row['T'])
                                        if isinstance(key, str):
                                            result_dict['2_time_' + key] = row['C']
                                        elif isinstance(key, tuple):
                                            _k, _p = key
                                            _p = _p['times']
                                            if row.get('P', 0) == _p:
                                                result_dict['2_time_' + _k] = row['C']

            df_data_dict['Ссылка'] = f'ID матча: {page_id}'
            df_data_dict['Страна'] = country_name
            df_data_dict['Лига'] = league_name
            df_data_dict['Команда 1'] = name_players[0]
            df_data_dict['Команда 2'] = name_players[1]
            df_data_dict['Дата'] = date_game if date_game else None
            df_data_dict['1'] = result_dict['P1']
            df_data_dict['Х'] = result_dict['X']
            df_data_dict['2'] = result_dict['P2']
            df_data_dict['1Х'] = result_dict['_1X']
            df_data_dict['12'] = result_dict['_12']
            df_data_dict['Х2'] = result_dict['_2X']
            df_data_dict['Ф1(0)'] = result_dict['F1_0']
            df_data_dict['Ф2(0)'] = result_dict['F2_0']
            df_data_dict['ТМ(2.5)'] = result_dict['TM']
            df_data_dict['ТБ(2.5)'] = result_dict['TB']
            df_data_dict['ИТМ1(1.5)'] = result_dict['IT1_men']
            df_data_dict['ИТБ1(1.5)'] = result_dict['IT1_bol']
            df_data_dict['ИТМ2(1.5)'] = result_dict['IT2_men']
            df_data_dict['ИТБ2(1.5)'] = result_dict['IT2_bol']
            df_data_dict['ОЗ Да'] = result_dict['ALL_win_yes']
            df_data_dict['ОЗ Нет'] = result_dict['ALL_win_no']
            df_data_dict['Гол оба тайма Да'] = result_dict['ALL_times_yes']
            df_data_dict['Гол оба тайма Нет'] = result_dict['ALL_times_no']

            df_data_dict['_1_1'] = result_dict['1_time_P1']
            df_data_dict['_1_Х'] = result_dict['1_time_X']
            df_data_dict['_1_2'] = result_dict['1_time_P2']
            df_data_dict['_1_1Х'] = result_dict['1_time__1X']
            df_data_dict['_1_12'] = result_dict['1_time__12']
            df_data_dict['_1_Х2'] = result_dict['1_time__2X']
            df_data_dict['_1_Ф1(0)'] = result_dict['1_time_F1_0']
            df_data_dict['_1_Ф2(0)'] = result_dict['1_time_F2_0']
            df_data_dict['_1_ТМ(1.5)'] = result_dict['1_time_TM']
            df_data_dict['_1_ТБ(1.5)'] = result_dict['1_time_TB']
            df_data_dict['_1_ИТМ1(0.5)'] = result_dict['1_time_IT1_men']
            df_data_dict['_1_ИТБ1(0.5)'] = result_dict['1_time_IT1_bol']
            df_data_dict['_1_ИТМ2(0.5)'] = result_dict['1_time_IT2_men']
            df_data_dict['_1_ИТБ2(0.5)'] = result_dict['1_time_IT2_bol']

            df_data_dict['_2_1'] = result_dict['2_time_P1']
            df_data_dict['_2_Х'] = result_dict['2_time_X']
            df_data_dict['_2_2'] = result_dict['2_time_P2']
            df_data_dict['_2_1Х'] = result_dict['2_time__1X']
            df_data_dict['_2_12'] = result_dict['2_time__12']
            df_data_dict['_2_Х2'] = result_dict['2_time__2X']
            df_data_dict['_2_Ф1(0)'] = result_dict['2_time_F1_0']
            df_data_dict['_2_Ф2(0)'] = result_dict['2_time_F2_0']
            df_data_dict['_2_ТМ(1.5)'] = result_dict['2_time_TM']
            df_data_dict['_2_ТБ(1.5)'] = result_dict['2_time_TB']
            df_data_dict['_2_ИТМ1(0.5)'] = result_dict['2_time_IT1_men']
            df_data_dict['_2_ИТБ1(0.5)'] = result_dict['2_time_IT1_bol']
            df_data_dict['_2_ИТМ2(0.5)'] = result_dict['2_time_IT2_men']
            df_data_dict['_2_ИТБ2(0.5)'] = result_dict['2_time_IT2_bol']

        return df_data_dict

    @classmethod
    def get_page_id(cls, page_link):
        result = None
        id_suffix = page_link.rsplit('/', 1)[-1]
        pattern = re.compile(r'(?P<id>^\d+).*$')
        id_result = pattern.search(id_suffix)
        if id_result:
            result = id_result.group('id')
        return result

    async def parse(self, browser):
        result = None
        msg = f'Открываем {self.url}'
        logger.info(msg)
        self.status = msg
        min_offset_values = {
            'За всё время': None,
            'Ближайшие 24 часа': 24 * 60,
            'Ближайшие 12 часов': 12 * 60,
            'Ближайшие 6 часов': 6 * 60,
            'Ближайшие 2 часа': 2 * 60,
            'Ближайший час': 60,
        }
        min_offset = min_offset_values.get(self.radio_period)
        ids = await self.get_all_ids(min_offset)
        df_data = []
        logger.info(f'Количество ссылок: {len(ids)}')
        self.count_links = len(ids)
        self.status = 'Собираем данные по каждому матчу'
        for page_id in self.tqdm(ids):
            df_data_dict = dict()
            attempt = 1
            while attempt < 3:
                try:
                    df_data_dict = await self._parse(
                        page_id
                    )
                    df_data.append(
                        df_data_dict
                    )
                except Exception:
                    attempt += 1
                    logger.exception('Ошибка')
                    await asyncio.sleep(5)
                else:
                    break
        await browser.close()
        if df_data:
            msg = f'Собрано данных: {len(df_data)}'
            logger.info(msg)
            self.status = msg
            df = pd.DataFrame.from_records(df_data)
            now_msk = datetime.now(tz=pytz.timezone('Europe/Moscow'))
            df['Дата слепка, МСК'] = now_msk
            columns = [
                'Ссылка',
                'Страна',
                'Лига',
                'Команда 1',
                'Команда 2',
                'Дата',
                'Дата слепка, МСК',
                '1',
                'Х',
                '2',
                '1Х',
                '12',
                'Х2',
                'Ф1(0)',
                'Ф2(0)',
                'ТМ(2.5)',
                'ТБ(2.5)',
                'ИТМ1(1.5)',
                'ИТБ1(1.5)',
                'ИТМ2(1.5)',
                'ИТБ2(1.5)',
                'ОЗ Да',
                'ОЗ Нет',
                'Гол оба тайма Да',
                'Гол оба тайма Нет',
                '_1_1',
                '_1_Х',
                '_1_2',
                '_1_1Х',
                '_1_12',
                '_1_Х2',
                '_1_Ф1(0)',
                '_1_Ф2(0)',
                '_1_ТМ(1.5)',
                '_1_ТБ(1.5)',
                '_1_ИТМ1(0.5)',
                '_1_ИТБ1(0.5)',
                '_1_ИТМ2(0.5)',
                '_1_ИТБ2(0.5)',
                '_2_1',
                '_2_Х',
                '_2_2',
                '_2_1Х',
                '_2_12',
                '_2_Х2',
                '_2_Ф1(0)',
                '_2_Ф2(0)',
                '_2_ТМ(1.5)',
                '_2_ТБ(1.5)',
                '_2_ИТМ1(0.5)',
                '_2_ИТБ1(0.5)',
                '_2_ИТМ2(0.5)',
                '_2_ИТБ2(0.5)',
            ]
            df = df.reindex(columns=columns)
            value_columns_start = columns.index('1')
            # решаем проблему округления числа 1.285 в 1.29, а не 1.28 путем прибавления 0.0001
            df.iloc[:, value_columns_start:] = (
                df.iloc[:, value_columns_start:].astype(np.float64) + pow(10, -4)
            ).round(2)
            df['Дата'] = df['Дата'].dt.tz_localize(None)
            df['Дата слепка, МСК'] = df['Дата слепка, МСК'].dt.tz_localize(None)
            older_data = Path('storage') / Path('older.json')
            older_df = pd.DataFrame(columns=columns)
            if older_data.exists() and not DEBUG:
                older_df = pd.read_json(older_data, date_unit='s')
                older_df['Дата'] = older_df['Дата'].astype('datetime64[s]')
                older_df['Дата слепка, МСК'] = older_df['Дата слепка, МСК'].astype('datetime64[s]')
            path = f'files/xbet_{now_msk.isoformat()}.xlsx'
            if older_df.empty:
                full_df = df
            else:
                full_df = pd.concat((df, older_df))
            full_df.to_excel('files/debug.xlsx', index=False, columns=columns)
            full_df = full_df.sort_values(
                [
                    'Дата',
                    'Команда 1',
                    'Команда 2',
                ],
                ascending=[False, True, True]
            )
            full_df['Double'] = full_df[['Команда 1', 'Команда 2']].duplicated()
            full_df = full_df.reset_index(drop=True)
            data = np.array(full_df[full_df['Double']].index.values)
            ddiff = np.diff(data)
            subArrays = np.split(data, np.where(ddiff != 1)[0]+1)

            with pd.ExcelWriter(path, datetime_format='%d.%m.%y %H:%M') as writer:
                full_df.to_excel(writer, index=False, startrow=1, columns=columns)
                workbook = writer.book

                sheet = workbook.active
                for i in range(1, sheet.max_column + 1):
                    sheet.cell(2, i).alignment = Alignment(text_rotation=90)

                match_index_start = columns.index('1') + 1
                first_time_index_start = columns.index('_1_1') + 1
                second_time_index_start = columns.index('_2_1') + 1

                match_index_end = first_time_index_start - 1
                first_time_index_end = second_time_index_start - 1
                second_time_index_end = len(columns)

                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                sheet.merge_cells(
                    start_row=1,
                    end_row=1,
                    start_column=match_index_start,
                    end_column=match_index_end
                )
                sheet.cell(1, match_index_start).value = 'Матч'
                sheet.cell(1, match_index_start).alignment = Alignment(horizontal='center')
                sheet.cell(1, match_index_start).border = thin_border

                sheet.merge_cells(
                    start_row=1,
                    end_row=1,
                    start_column=first_time_index_start,
                    end_column=first_time_index_end
                )
                sheet.cell(1, first_time_index_start).value = '1 тайм'
                sheet.cell(1, first_time_index_start).alignment = Alignment(horizontal='center')
                sheet.cell(1, first_time_index_start).border = thin_border

                sheet.merge_cells(
                    start_row=1,
                    end_row=1,
                    start_column=second_time_index_start,
                    end_column=second_time_index_end
                )
                sheet.cell(1, second_time_index_start).value = '2 тайм'
                sheet.cell(1, second_time_index_start).alignment = Alignment(horizontal='center')
                sheet.cell(1, second_time_index_start).border = thin_border

                for i in range(first_time_index_start, first_time_index_end + 2):
                    sheet.cell(2, i - 1).value = sheet.cell(2, i - 1).value.replace('_1_', '')

                for i in range(second_time_index_start, second_time_index_end + 2):
                    sheet.cell(2, i - 1).value = sheet.cell(2, i - 1).value.replace('_2_', '')

                for subArray in subArrays:
                    if subArray.size > 0:
                        sheet.row_dimensions.group(subArray[0] + 3, subArray[-1] + 3, hidden=True)

                workbook.save(path)
            full_df.to_json(older_data, date_unit='s', date_format='iso')
            result = FileResponse(
                path=path,
                filename=path
            )
        else:
            return PlainTextResponse('Не собрали данных')
        return result
