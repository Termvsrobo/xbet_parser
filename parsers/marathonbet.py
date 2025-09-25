import asyncio
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from urllib.parse import urlparse

import numpy as np
import pandas as pd
import pytz
from bs4 import BeautifulSoup
from fastapi.responses import FileResponse, PlainTextResponse
from loguru import logger
from openpyxl.styles import Alignment, Border, Side
from playwright._impl._errors import Error, TimeoutError

from base import Parser
from utils import parse_date_str

logger.add('logs/marathonbet.log')

DEBUG = True


def get_players_links(page_content):
    soup = BeautifulSoup(page_content, 'html.parser')
    tables = soup.find_all(
        lambda tag: tag.name == 'table' and tag.get('class') == ['coupon-row-item']
    )
    players_links = set()
    for table in tables:
        name_column = table.find('td', attrs={'class': 'first'})
        players = name_column.find_all('a', attrs={'class': 'member-link'})
        for i, player in enumerate(players, 1):
            players_links.add(player.attrs.get('href'))
    return players_links


def parse(page_content, page_link):
    df_data_dict = dict()
    soup = BeautifulSoup(page_content, 'html.parser')
    country_name = None
    league_name = None
    league_header = soup.find(lambda tag: tag.name == 'h2' and tag.get('class') == ['category-label'])
    if league_header:
        league_header_data = [
            child.text
            for child in league_header.find_all(
                lambda tag: tag.name == 'span' and tag.get('class') == ['nowrap']
            )
        ]
        if len(league_header_data) == 2:
            country_name = league_header_data[0]
            league_name = league_header_data[1]
        elif len(league_header_data) > 2:
            country_name = league_header_data[0]
            league_name = ' '.join(league_header_data[1:])
    tables = soup.find_all(lambda tag: tag.name == 'table' and tag.get('class') == ['coupon-row-item'])
    for table in tables:
        name_column = table.find('td', attrs={'class': 'first'})
        players = name_column.find_all('a', attrs={'class': 'member-link'})
        players_names = []
        for i, player in enumerate(players, 1):
            players_names.append(f'{i}. {player.text.replace("\n", "")}')
    name = '\n'.join(players_names)

    name_players = name.split('\n')
    name_players = [name_players[0].replace('1. ', ''), name_players[1].replace('2. ', '')]
    if table.find('div', attrs={'class': 'date-wrapper'}):
        date_game = table.find('div', attrs={'class': 'date-wrapper'}).text
    else:
        date_game = None
    _results = soup.find(
        lambda tag: tag.name == 'div' and tag.get('class') == ['name-field'] and tag.text.strip() == 'Результат'
    )
    if _results:
        results = _results.find_parent(lambda tag: tag.name == 'div' and tag.get('class') == ['block-market-wrapper'])
    else:
        results = None

    _head_starts = soup.find(
        lambda tag: tag.name == 'div' and tag.get('class') == ['name-field'] and tag.text.strip() == 'Форы'
    )
    if _head_starts:
        head_starts = _head_starts.find_parent(
            lambda tag: tag.name == 'div' and tag.get('class') == ['block-market-wrapper']
        )
    else:
        head_starts = None

    _totals = soup.find(
        lambda tag: tag.name == 'div' and tag.get('class') == ['name-field'] and tag.text.strip() == 'Тоталы'
    )
    if _totals:
        totals = _totals.find_parent(lambda tag: tag.name == 'div' and tag.get('class') == ['block-market-wrapper'])
    else:
        totals = None

    _goals = soup.find(
        lambda tag: tag.name == 'div' and tag.get('class') == ['name-field'] and tag.text.strip() == 'Голы'
    )
    if _goals:
        goals = _goals.find_parent(lambda tag: tag.name == 'div' and tag.get('class') == ['block-market-wrapper'])
    else:
        goals = None

    _times = soup.find(
        lambda tag: tag.name == 'div' and tag.get('class') == ['name-field'] and tag.text.strip() == 'Таймы'
    )
    if _times:
        times = _times.find_parent(lambda tag: tag.name == 'div' and tag.get('class') == ['block-market-wrapper'])
    else:
        times = None

    # Результаты
    result_dict = defaultdict(lambda: None)
    if results:
        for key, value in (
            ('P1', f'{name_players[0]} (победа)'),
            ('X', 'Ничья'),
            ('P2', f'{name_players[1]} (победа)'),
            ('_1X', f'{name_players[0]} (победа) или ничья'),
            ('_12', f'{name_players[0]} (победа) или {name_players[1]} (победа)'),
            ('_2X', f'{name_players[1]} (победа) или ничья'),
        ):
            element = results.find(
                lambda tag: tag.name == 'div' and tag.get('class') == ['result-left'] and tag.text.strip() == value
            )
            if element:
                result_dict[key] = element.find_next_sibling(
                    lambda tag: tag.name == 'div' and tag.get('class') == ['result-right']
                ).span.text

    # Форы
    head_starts_dict = defaultdict(lambda: None)
    if head_starts:
        head_starts_visible_div = head_starts.find(
            lambda tag: tag.name == 'div' and tag.get('class') == ['market-inline-block-table-wrapper']
        )
        head_starts_table = head_starts_visible_div.find('table', class_='td-border')
        table_data = [t.find_all('td') for t in head_starts_table.tbody.find_all('tr')]
        left_table_data = [t[0] for t in table_data if t]
        right_table_data = [t[1] for t in table_data if t]
        for key, table in (
            ('F1_0', left_table_data),
            ('F2_0', right_table_data),
        ):
            for td in table:
                _td = td.find(
                    lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(0)' in tag.text
                )
                if _td:
                    head_starts_dict[key] = _td.find_next_sibling().span.text
                    break

    # Тотал голов
    totals_dict = defaultdict(lambda: None)
    if totals:
        totals_visible_div = totals.find_all(
            lambda tag: tag.name == 'div' and tag.get('class') == ['market-inline-block-table-wrapper']
        )
        totals_table_div = next(
            filter(lambda x: 'Тотал голов' in x.text, totals_visible_div),
            None
        )
        it_1_totals_table_div = next(
            filter(lambda x: f'Тотал голов ({name_players[0]})' in x.text, totals_visible_div),
            None
        )
        it_2_totals_table_div = next(
            filter(lambda x: f'Тотал голов ({name_players[1]})' in x.text, totals_visible_div),
            None
        )

        if totals_table_div:
            totals_table = totals_table_div.find('table', class_='td-border')
            table_data = [t.find_all('td') for t in totals_table.tbody.find_all('tr')]
            left_table_data = [t[0] for t in table_data if t]
            right_table_data = [t[1] for t in table_data if t]
            for key, table in (
                ('TM_25', left_table_data),
                ('TB_25', right_table_data),
            ):
                for td in table:
                    _td = td.find(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(2.5)' in tag.text
                    )
                    if _td:
                        totals_dict[key] = _td.find_next_sibling().span.text
                        break

        if it_1_totals_table_div:
            it_1_totals_table = it_1_totals_table_div.find('table', class_='td-border')
            table_data = [t.find_all('td') for t in it_1_totals_table.tbody.find_all('tr')]
            left_table_data = [t[0] for t in table_data if t]
            right_table_data = [t[1] for t in table_data if t]
            for key, table in (
                ('IT1_men_15', left_table_data),
                ('IT1_bol_15', right_table_data),
            ):
                for td in table:
                    _td = td.find(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(1.5)' in tag.text
                    )
                    if _td:
                        totals_dict[key] = _td.find_next_sibling().span.text
                        break

        if it_2_totals_table_div:
            it_2_totals_table = it_2_totals_table_div.find('table', class_='td-border')
            table_data = [t.find_all('td') for t in it_2_totals_table.tbody.find_all('tr')]
            left_table_data = [t[0] for t in table_data if t]
            right_table_data = [t[1] for t in table_data if t]
            for key, table in (
                ('IT2_men_15', left_table_data),
                ('IT2_bol_15', right_table_data),
            ):
                for td in table:
                    _td = td.find(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(1.5)' in tag.text
                    )
                    if _td:
                        totals_dict[key] = _td.find_next_sibling().span.text
                        break

    # Голы
    goals_dict = defaultdict(lambda: None)
    if goals:
        yes_column_exist = bool(goals.find(
            lambda tag: tag.name == 'th' and tag.get('class') == ['width25'] and tag.span.text == 'Да'
        ))
        no_column_exist = bool(goals.find(
            lambda tag: tag.name == 'th' and tag.get('class') == ['width25'] and tag.span.text == 'Нет'
        ))
        _all_win = goals.find(lambda tag: tag.text == 'Обе команды забьют')
        if _all_win:
            ALL_win = _all_win.parent
            if ALL_win:
                ALL_win_yes_no = ALL_win.parent.find_all(
                    lambda tag: tag.name == 'td' and sorted(tag.get('class')) == sorted(
                        ['price', 'height-column-with-price']
                    )
                )
                if yes_column_exist and no_column_exist:
                    goals_dict['ALL_win_yes'], goals_dict['ALL_win_no'] = [
                        all_win_yes_no.span.text
                        for all_win_yes_no in ALL_win_yes_no
                    ]
                elif yes_column_exist:
                    goals_dict['ALL_win_yes'] = ALL_win_yes_no[0].span.text
                elif no_column_exist:
                    goals_dict['ALL_win_no'] = ALL_win_yes_no[0].span.text

        _all_times = goals.find(lambda tag: tag.text == 'Голы в обоих таймах')
        if _all_times:
            ALL_times = _all_times.parent
            if ALL_times:
                ALL_times_yes_no = ALL_times.parent.find_all(
                    lambda tag: tag.name == 'td' and sorted(tag.get('class')) == sorted(
                        ['price', 'height-column-with-price']
                    )
                )
                if yes_column_exist and no_column_exist:
                    goals_dict['ALL_times_yes'], goals_dict['ALL_times_no'] = [
                        all_times_yes_no.span.text
                        for all_times_yes_no in ALL_times_yes_no
                    ]
                elif yes_column_exist:
                    goals_dict['ALL_times_yes'] = ALL_win_yes_no[0].span.text
                elif no_column_exist:
                    goals_dict['ALL_times_no'] = ALL_win_yes_no[0].span.text

        _goals_it1_1_time = goals.find(lambda tag: tag.text == f'{name_players[0]} забьет, 1-й тайм')
        if _goals_it1_1_time:
            _it1_1_time = _goals_it1_1_time.parent
            if _it1_1_time:
                _it1_1_time_data = _it1_1_time.parent.find_all(
                    lambda tag: tag.name == 'td' and sorted(tag.get('class')) == sorted(
                        ['price', 'height-column-with-price']
                    )
                )
                goals_dict['IT1_bol_05_1_time'], goals_dict['IT1_men_05_1_time'] = [
                    _t.span.text
                    for _t in _it1_1_time_data
                ]

        _goals_it1_2_time = goals.find(lambda tag: tag.text == f'{name_players[0]} забьет, 2-й тайм')
        if _goals_it1_2_time:
            _it1_2_time = _goals_it1_2_time.parent
            if _it1_2_time:
                _it1_2_time_data = _it1_2_time.parent.find_all(
                    lambda tag: tag.name == 'td' and sorted(tag.get('class')) == sorted(
                        ['price', 'height-column-with-price']
                    )
                )
                goals_dict['IT1_bol_05_2_time'], goals_dict['IT1_men_05_2_time'] = [
                    _t.span.text
                    for _t in _it1_2_time_data
                ]

        _goals_it2_1_time = goals.find(lambda tag: tag.text == f'{name_players[1]} забьет, 1-й тайм')
        if _goals_it2_1_time:
            _it2_1_time = _goals_it2_1_time.parent
            if _it2_1_time:
                _it2_1_time_data = _it2_1_time.parent.find_all(
                    lambda tag: tag.name == 'td' and sorted(tag.get('class')) == sorted(
                        ['price', 'height-column-with-price']
                    )
                )
                goals_dict['IT2_bol_05_1_time'], goals_dict['IT2_men_05_1_time'] = [
                    _t.span.text
                    for _t in _it2_1_time_data
                ]

        _goals_it2_2_time = goals.find(lambda tag: tag.text == f'{name_players[1]} забьет, 2-й тайм')
        if _goals_it2_2_time:
            _it2_2_time = _goals_it2_2_time.parent
            if _it2_2_time:
                _it2_2_time_data = _it2_2_time.parent.find_all(
                    lambda tag: tag.name == 'td' and sorted(tag.get('class')) == sorted(
                        ['price', 'height-column-with-price']
                    )
                )
                goals_dict['IT2_bol_05_2_time'], goals_dict['IT2_men_05_2_time'] = [
                    _t.span.text
                    for _t in _it2_2_time_data
                ]

    # Таймы
    times_dict = defaultdict(lambda: None)
    if times:
        times_elements = times.find_all(
            lambda tag: tag.name == 'div' and tag.get('class') == ['market-inline-block-table-wrapper']
        )

        results_1_time = next(
            filter(lambda x: 'Результат, 1-й тайм' in x.text, times_elements),
            None
        )
        win_head_start_1_time = next(
            filter(lambda x: 'Победа с учетом форы, 1-й тайм' in x.text, times_elements),
            None
        )
        goals_1_time = next(
            filter(lambda x: 'Тотал голов, 1-й тайм' in x.text, times_elements),
            None
        )
        goals_it1_1_time = next(
            filter(lambda x: f'Тотал голов ({name_players[0]}), 1-й тайм' in x.text, times_elements),
            None
        )
        goals_it2_1_time = next(
            filter(lambda x: f'Тотал голов ({name_players[1]}), 1-й тайм' in x.text, times_elements),
            None
        )
        results_2_time = next(
            filter(lambda x: 'Результат, 2-й тайм' in x.text, times_elements),
            None
        )
        win_head_start_2_time = next(
            filter(lambda x: 'Победа с учетом форы, 2-й тайм' in x.text, times_elements),
            None
        )
        goals_2_time = next(
            filter(lambda x: 'Тотал голов, 2-й тайм' in x.text, times_elements),
            None
        )
        goals_it1_2_time = next(
            filter(lambda x: f'Тотал голов ({name_players[0]}), 2-й тайм' in x.text, times_elements),
            None
        )
        goals_it2_2_time = next(
            filter(lambda x: f'Тотал голов ({name_players[1]}), 2-й тайм' in x.text, times_elements),
            None
        )

        # --
        if results_1_time:
            for key, value in (
                ('goal_1_time_P1', f'{name_players[0]} (победа)'),
                ('goal_1_time_X', 'Ничья'),
                ('goal_1_time_P2', f'{name_players[1]} (победа)'),
                ('goal_1_time_1X', f'{name_players[0]} (победа) или ничья'),
                ('goal_1_time_12', f'{name_players[0]} (победа) или {name_players[1]} (победа)'),
                ('goal_1_time_2X', f'{name_players[1]} (победа) или ничья'),
            ):
                element = results_1_time.find(
                    lambda tag: tag.name == 'div' and tag.get('class') == ['result-left'] and tag.text.strip() == value
                )
                if element:
                    times_dict[key] = element.find_next_sibling(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['result-right']
                    ).span.text

        # --
        if win_head_start_1_time:
            head_starts_table_1_time = win_head_start_1_time.find('table', class_='td-border')
            table_data = [t.find_all('td') for t in head_starts_table_1_time.tbody.find_all('tr')]
            left_table_data = [t[0] for t in table_data if t]
            right_table_data = [t[1] for t in table_data if t]
            for key, table in (
                ('F1_0_1_time', left_table_data),
                ('F2_0_1_time', right_table_data),
            ):
                for td in table:
                    _td = td.find(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(0)' in tag.text
                    )
                    if _td:
                        times_dict[key] = _td.find_next_sibling().span.text
                        break

        # --
        if goals_1_time:
            totals_table_1_time = goals_1_time.find('table', class_='td-border')
            table_data = [t.find_all('td') for t in totals_table_1_time.tbody.find_all('tr')]
            left_table_data = [t[0] for t in table_data if t]
            right_table_data = [t[1] for t in table_data if t]
            for key, table in (
                ('TM_15_1_time', left_table_data),
                ('TB_15_1_time', right_table_data),
            ):
                for td in table:
                    _td = td.find(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(1.5)' in tag.text
                    )
                    if _td:
                        times_dict[key] = _td.find_next_sibling().span.text
                        break

        # --
        if goals_it1_1_time:
            it_1_totals_table_1_time = goals_it1_1_time.find('table', class_='td-border')
            table_data = [t.find_all('td') for t in it_1_totals_table_1_time.tbody.find_all('tr')]
            left_table_data = [t[0] for t in table_data if t]
            right_table_data = [t[1] for t in table_data if t]
            for key, table in (
                ('IT1_men_05_1_time', left_table_data),
                ('IT1_bol_05_1_time', right_table_data),
            ):
                for td in table:
                    _td = td.find(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(0.5)' in tag.text
                    )
                    if _td:
                        times_dict[key] = _td.find_next_sibling().span.text
                        break

        if goals_it2_1_time:
            it_2_totals_table_1_time = goals_it2_1_time.find('table', class_='td-border')
            table_data = [t.find_all('td') for t in it_2_totals_table_1_time.tbody.find_all('tr')]
            left_table_data = [t[0] for t in table_data if t]
            right_table_data = [t[1] for t in table_data if t]
            for key, table in (
                ('IT2_men_05_1_time', left_table_data),
                ('IT2_bol_05_1_time', right_table_data),
            ):
                for td in table:
                    _td = td.find(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(0.5)' in tag.text
                    )
                    if _td:
                        times_dict[key] = _td.find_next_sibling().span.text
                        break

        # 2 тайм
        # --
        if results_2_time:
            for key, value in (
                ('goal_2_time_P1', f'{name_players[0]} (победа)'),
                ('goal_2_time_X', 'Ничья'),
                ('goal_2_time_P2', f'{name_players[1]} (победа)'),
                ('goal_2_time_1X', f'{name_players[0]} (победа) или ничья'),
                ('goal_2_time_12', f'{name_players[0]} (победа) или {name_players[1]} (победа)'),
                ('goal_2_time_2X', f'{name_players[1]} (победа) или ничья'),
            ):
                element = results_2_time.find(
                    lambda tag: tag.name == 'div' and tag.get('class') == ['result-left'] and tag.text.strip() == value
                )
                if element:
                    times_dict[key] = element.find_next_sibling(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['result-right']
                    ).span.text

        # --
        if win_head_start_2_time:
            head_starts_table_2_time = win_head_start_2_time.find('table', class_='td-border')
            table_data = [t.find_all('td') for t in head_starts_table_2_time.tbody.find_all('tr')]
            left_table_data = [t[0] for t in table_data if t]
            right_table_data = [t[1] for t in table_data if t]
            for key, table in (
                ('F1_0_2_time', left_table_data),
                ('F2_0_2_time', right_table_data),
            ):
                for td in table:
                    _td = td.find(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(0)' in tag.text
                    )
                    if _td:
                        times_dict[key] = _td.find_next_sibling().span.text
                        break

        # --
        if goals_2_time:
            totals_table_2_time = goals_2_time.find('table', class_='td-border')
            table_data = [t.find_all('td') for t in totals_table_2_time.tbody.find_all('tr')]
            left_table_data = [t[0] for t in table_data if t]
            right_table_data = [t[1] for t in table_data if t]
            for key, table in (
                ('TM_15_2_time', left_table_data),
                ('TB_15_2_time', right_table_data),
            ):
                for td in table:
                    _td = td.find(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(1.5)' in tag.text
                    )
                    if _td:
                        times_dict[key] = _td.find_next_sibling().span.text
                        break

        if goals_it1_2_time:
            it_1_totals_table_2_time = goals_it1_2_time.find('table', class_='td-border')
            table_data = [t.find_all('td') for t in it_1_totals_table_2_time.tbody.find_all('tr')]
            left_table_data = [t[0] for t in table_data if t]
            right_table_data = [t[1] for t in table_data if t]
            for key, table in (
                ('IT1_men_05_2_time', left_table_data),
                ('IT1_bol_05_2_time', right_table_data),
            ):
                for td in table:
                    _td = td.find(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(0.5)' in tag.text
                    )
                    if _td:
                        times_dict[key] = _td.find_next_sibling().span.text
                        break

        if goals_it2_2_time:
            it_2_totals_table_2_time = goals_it2_2_time.find('table', class_='td-border')
            table_data = [t.find_all('td') for t in it_2_totals_table_2_time.tbody.find_all('tr')]
            left_table_data = [t[0] for t in table_data if t]
            right_table_data = [t[1] for t in table_data if t]
            for key, table in (
                ('IT2_men_05_2_time', left_table_data),
                ('IT2_bol_05_2_time', right_table_data),
            ):
                for td in table:
                    _td = td.find(
                        lambda tag: tag.name == 'div' and tag.get('class') == ['coeff-value'] and '(0.5)' in tag.text
                    )
                    if _td:
                        times_dict[key] = _td.find_next_sibling().span.text
                        break

    for k in (
        'IT1_men_05_1_time',
        'IT1_bol_05_1_time',
        'IT1_men_05_2_time',
        'IT1_bol_05_2_time',
        'IT2_men_05_1_time',
        'IT2_bol_05_1_time',
        'IT2_men_05_2_time',
        'IT2_bol_05_2_time',
    ):
        if not times_dict[k] and goals_dict[k]:
            times_dict[k] = goals_dict[k]

    df_data_dict['Ссылка'] = page_link
    df_data_dict['Страна'] = country_name
    df_data_dict['Лига'] = league_name
    df_data_dict['Команда 1'] = name_players[0]
    df_data_dict['Команда 2'] = name_players[1]
    df_data_dict['Дата'] = parse_date_str(date_game) if date_game else None
    df_data_dict['1'] = result_dict['P1']
    df_data_dict['Х'] = result_dict['X']
    df_data_dict['2'] = result_dict['P2']
    df_data_dict['1Х'] = result_dict['_1X']
    df_data_dict['12'] = result_dict['_12']
    df_data_dict['Х2'] = result_dict['_2X']
    df_data_dict['Ф1(0)'] = head_starts_dict['F1_0']
    df_data_dict['Ф2(0)'] = head_starts_dict['F2_0']
    df_data_dict['ТМ(2.5)'] = totals_dict['TM_25']
    df_data_dict['ТБ(2.5)'] = totals_dict['TB_25']
    df_data_dict['ИТМ1(1.5)'] = totals_dict['IT1_men_15']
    df_data_dict['ИТБ1(1.5)'] = totals_dict['IT1_bol_15']
    df_data_dict['ИТМ2(1.5)'] = totals_dict['IT2_men_15']
    df_data_dict['ИТБ2(1.5)'] = totals_dict['IT2_bol_15']
    df_data_dict['ОЗ Да'] = goals_dict['ALL_win_yes']
    df_data_dict['ОЗ Нет'] = goals_dict['ALL_win_no']
    df_data_dict['Гол оба тайма Да'] = goals_dict['ALL_times_yes']
    df_data_dict['Гол оба тайма Нет'] = goals_dict['ALL_times_no']

    df_data_dict['_1_1'] = times_dict['goal_1_time_P1']
    df_data_dict['_1_Х'] = times_dict['goal_1_time_X']
    df_data_dict['_1_2'] = times_dict['goal_1_time_P2']
    df_data_dict['_1_1Х'] = times_dict['goal_1_time_1X']
    df_data_dict['_1_12'] = times_dict['goal_1_time_12']
    df_data_dict['_1_Х2'] = times_dict['goal_1_time_2X']
    df_data_dict['_1_Ф1(0)'] = times_dict['F1_0_1_time']
    df_data_dict['_1_Ф2(0)'] = times_dict['F2_0_1_time']
    df_data_dict['_1_ТМ(1.5)'] = times_dict['TM_15_1_time']
    df_data_dict['_1_ТБ(1.5)'] = times_dict['TB_15_1_time']
    df_data_dict['_1_ИТМ1(0.5)'] = times_dict['IT1_men_05_1_time']
    df_data_dict['_1_ИТБ1(0.5)'] = times_dict['IT1_bol_05_1_time']
    df_data_dict['_1_ИТМ2(0.5)'] = times_dict['IT2_men_05_1_time']
    df_data_dict['_1_ИТБ2(0.5)'] = times_dict['IT2_bol_05_1_time']

    df_data_dict['_2_1'] = times_dict['goal_2_time_P1']
    df_data_dict['_2_Х'] = times_dict['goal_2_time_X']
    df_data_dict['_2_2'] = times_dict['goal_2_time_P2']
    df_data_dict['_2_1Х'] = times_dict['goal_2_time_1X']
    df_data_dict['_2_12'] = times_dict['goal_2_time_12']
    df_data_dict['_2_Х2'] = times_dict['goal_2_time_2X']
    df_data_dict['_2_Ф1(0)'] = times_dict['F1_0_2_time']
    df_data_dict['_2_Ф2(0)'] = times_dict['F2_0_2_time']
    df_data_dict['_2_ТМ(1.5)'] = times_dict['TM_15_2_time']
    df_data_dict['_2_ТБ(1.5)'] = times_dict['TB_15_2_time']
    df_data_dict['_2_ИТМ1(0.5)'] = times_dict['IT1_men_05_2_time']
    df_data_dict['_2_ИТБ1(0.5)'] = times_dict['IT1_bol_05_2_time']
    df_data_dict['_2_ИТМ2(0.5)'] = times_dict['IT2_men_05_2_time']
    df_data_dict['_2_ИТБ2(0.5)'] = times_dict['IT2_bol_05_2_time']

    return df_data_dict


class MarathonbetParser(Parser):
    async def parse(self, browser):
        result = None
        msg = f'Открываем {self.url}'
        logger.info(msg)
        self.status = msg
        page = await browser.new_page()
        page.set_default_timeout(180000)
        await page.goto('/')
        await page.wait_for_load_state()
        await page.goto('su')
        await page.wait_for_load_state()
        msg = 'Ждем окончания проверки браузера'
        logger.info(msg)
        self.status = msg
        try:
            while 'Just' in await page.title():
                await asyncio.sleep(1)
        except Error as exc:
            logger.exception(exc.message)
            page = browser.pages[-1]
            return PlainTextResponse('Во время обработки произошла ошибка. Попробуйте позже.')
        else:
            if urlparse(page.url).path != '/su/':
                await page.goto('su')
                await page.wait_for_load_state()
            try:
                await page.wait_for_selector(
                    '//table[@class="coupon-row-item"]',
                    timeout=180000
                )
            except TimeoutError as exc:
                logger.exception(exc.message)
                return PlainTextResponse('Вышло время ожидания страницы. Попробуйте позже.')
            else:
                msg = f'Собираем список матчей по футболу за {self.radio_period}'
                logger.info(msg)
                self.status = msg
                await page.get_by_text('Футбол').first.click()
                await page.get_by_text(self.radio_period).first.click()
                need_scroll = True
                attempts = 5
                content = ''
                while need_scroll:
                    await page.mouse.wheel(0, 3000)
                    await page.wait_for_load_state()
                    await page.wait_for_timeout(2000)
                    if content == await page.content():
                        if attempts:
                            attempts -= 1
                        else:
                            need_scroll = False
                    else:
                        content = await page.content()
                df_data = []
                players_links = get_players_links(await page.content())
                logger.info(f'Количество ссылок: {len(players_links)}')
                self.count_links = len(players_links)
                self.status = 'Собираем данные по каждому матчу'
                for player_link in self.tqdm(players_links):
                    df_data_dict = dict()
                    attempt = 1
                    while attempt < 3:
                        try:
                            player_page = await browser.new_page()
                            await player_page.goto(player_link)
                            await player_page.wait_for_load_state()
                            await player_page.wait_for_selector(
                                '//div[@class="block-market-wrapper"]',
                                timeout=180000
                            )
                            df_data_dict = parse(
                                await player_page.content(),
                                self.url + player_link[1:]
                            )
                            await player_page.close()
                            df_data.append(
                                df_data_dict
                            )
                        except Exception:
                            attempt += 1
                            await asyncio.sleep(5)
                        else:
                            break
                await browser.close()
                if df_data:
                    msg = f'Собрано данных: {len(df_data)}'
                    logger.info(msg)
                    self.status = msg
                    df = pd.DataFrame.from_records(df_data)
                    now_msk = datetime.now(tz=pytz.timezone('Europe/Moscow'))
                    df['Дата слепка, МСК'] = now_msk
                    columns = [
                        'Ссылка',
                        'Страна',
                        'Лига',
                        'Команда 1',
                        'Команда 2',
                        'Дата',
                        'Дата слепка, МСК',
                        '1',
                        'Х',
                        '2',
                        '1Х',
                        '12',
                        'Х2',
                        'Ф1(0)',
                        'Ф2(0)',
                        'ТМ(2.5)',
                        'ТБ(2.5)',
                        'ИТМ1(1.5)',
                        'ИТБ1(1.5)',
                        'ИТМ2(1.5)',
                        'ИТБ2(1.5)',
                        'ОЗ Да',
                        'ОЗ Нет',
                        'Гол оба тайма Да',
                        'Гол оба тайма Нет',
                        '_1_1',
                        '_1_Х',
                        '_1_2',
                        '_1_1Х',
                        '_1_12',
                        '_1_Х2',
                        '_1_Ф1(0)',
                        '_1_Ф2(0)',
                        '_1_ТМ(1.5)',
                        '_1_ТБ(1.5)',
                        '_1_ИТМ1(0.5)',
                        '_1_ИТБ1(0.5)',
                        '_1_ИТМ2(0.5)',
                        '_1_ИТБ2(0.5)',
                        '_2_1',
                        '_2_Х',
                        '_2_2',
                        '_2_1Х',
                        '_2_12',
                        '_2_Х2',
                        '_2_Ф1(0)',
                        '_2_Ф2(0)',
                        '_2_ТМ(1.5)',
                        '_2_ТБ(1.5)',
                        '_2_ИТМ1(0.5)',
                        '_2_ИТБ1(0.5)',
                        '_2_ИТМ2(0.5)',
                        '_2_ИТБ2(0.5)',
                    ]
                    df = df.reindex(columns=columns)
                    value_columns_start = columns.index('1')
                    # решаем проблему округления числа 1.285 в 1.29, а не 1.28 путем прибавления 0.0001
                    df.iloc[:, value_columns_start:] = (
                        df.iloc[:, value_columns_start:].astype(np.float64) + pow(10, -4)
                    ).round(2)
                    df['Дата'] = df['Дата'].dt.tz_localize(None)
                    df['Дата слепка, МСК'] = df['Дата слепка, МСК'].dt.tz_localize(None)
                    older_data = Path('storage') / Path('older.json')
                    older_df = pd.DataFrame(columns=columns)
                    if older_data.exists() and not DEBUG:
                        older_df = pd.read_json(older_data, date_unit='s')
                        older_df['Дата'] = older_df['Дата'].astype('datetime64[s]')
                        older_df['Дата слепка, МСК'] = older_df['Дата слепка, МСК'].astype('datetime64[s]')
                    path = f'files/marathonbet_{now_msk.isoformat()}.xlsx'
                    if older_df.empty:
                        full_df = df
                    else:
                        full_df = pd.concat((df, older_df))
                    full_df.to_excel('files/debug.xlsx', index=False, columns=columns)
                    full_df = full_df.sort_values(
                        [
                            'Дата',
                            'Команда 1',
                            'Команда 2',
                        ],
                        ascending=[False, True, True]
                    )
                    full_df['Double'] = full_df[['Команда 1', 'Команда 2']].duplicated()
                    full_df = full_df.reset_index(drop=True)
                    data = np.array(full_df[full_df['Double']].index.values)
                    ddiff = np.diff(data)
                    subArrays = np.split(data, np.where(ddiff != 1)[0]+1)

                    with pd.ExcelWriter(path, datetime_format='%d.%m.%y %H:%M') as writer:
                        full_df.to_excel(writer, index=False, startrow=1, columns=columns)
                        workbook = writer.book

                        sheet = workbook.active
                        for i in range(1, sheet.max_column + 1):
                            sheet.cell(2, i).alignment = Alignment(text_rotation=90)

                        match_index_start = columns.index('1') + 1
                        first_time_index_start = columns.index('_1_1') + 1
                        second_time_index_start = columns.index('_2_1') + 1

                        match_index_end = first_time_index_start - 1
                        first_time_index_end = second_time_index_start - 1
                        second_time_index_end = len(columns)

                        thin_border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )

                        sheet.merge_cells(
                            start_row=1,
                            end_row=1,
                            start_column=match_index_start,
                            end_column=match_index_end
                        )
                        sheet.cell(1, match_index_start).value = 'Матч'
                        sheet.cell(1, match_index_start).alignment = Alignment(horizontal='center')
                        sheet.cell(1, match_index_start).border = thin_border

                        sheet.merge_cells(
                            start_row=1,
                            end_row=1,
                            start_column=first_time_index_start,
                            end_column=first_time_index_end
                        )
                        sheet.cell(1, first_time_index_start).value = '1 тайм'
                        sheet.cell(1, first_time_index_start).alignment = Alignment(horizontal='center')
                        sheet.cell(1, first_time_index_start).border = thin_border

                        sheet.merge_cells(
                            start_row=1,
                            end_row=1,
                            start_column=second_time_index_start,
                            end_column=second_time_index_end
                        )
                        sheet.cell(1, second_time_index_start).value = '2 тайм'
                        sheet.cell(1, second_time_index_start).alignment = Alignment(horizontal='center')
                        sheet.cell(1, second_time_index_start).border = thin_border

                        for i in range(first_time_index_start, first_time_index_end + 2):
                            sheet.cell(2, i - 1).value = sheet.cell(2, i - 1).value.replace('_1_', '')

                        for i in range(second_time_index_start, second_time_index_end + 2):
                            sheet.cell(2, i - 1).value = sheet.cell(2, i - 1).value.replace('_2_', '')

                        for subArray in subArrays:
                            if subArray.size > 0:
                                sheet.row_dimensions.group(subArray[0] + 3, subArray[-1] + 3, hidden=True)

                        workbook.save(path)
                    full_df.to_json(older_data, date_unit='s', date_format='iso')
                    result = FileResponse(
                        path=path,
                        filename=path
                    )
                else:
                    return PlainTextResponse('Не собрали данных')
        return result
