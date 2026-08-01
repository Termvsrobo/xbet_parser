import gettext
import json
import re
from asyncio import to_thread
from collections import defaultdict
from contextlib import asynccontextmanager
from copy import copy
from datetime import datetime
from decimal import ROUND_DOWN, Decimal
from enum import IntEnum
from functools import partial
from itertools import count
from pathlib import Path
from typing import Annotated, ClassVar, Literal
from urllib.parse import parse_qs, unquote, urlencode, urljoin, urlparse, urlunparse

import httpx
import numpy as np
import pandas as pd
import pycountry as pc
from aiopath import AsyncPath
from bs4 import BeautifulSoup
from fastapi.responses import FileResponse, JSONResponse, PlainTextResponse
from openpyxl.styles import Border, Side
from openpyxl.worksheet.cell_range import CellRange
from pydantic import (
    BaseModel,
    Discriminator,
    Field,
    PositiveInt,
    RootModel,
    Tag,
    TypeAdapter,
)
from tqdm import tqdm
from xlsxtpl.writerx import BookWriter

from base import Parser
from config import settings


class FieldType(IntEnum):
    BOOL: int = 1
    TIME: int = 2
    FLOAT: int = 3
    STR: int = 4


def filter_type_discriminator(v):
    result = None
    if isinstance(v, dict):
        result = v.get('type', None)
    else:
        result = getattr(v, 'type', None)
    return result


class BaseFilterField(BaseModel):
    filter_value: str
    column: int
    priority: PositiveInt | None = None

    def get_value(self, value, filter_value: str | None = None):
        return value

    def next_value(self, value):
        _filter_value = self.filter_value
        for _ in range(2):
            yield self.get_value(value, _filter_value)
            _filter_value = _filter_value[:-1]

    class Config:
        validate_assignment = True


class FloatField(BaseFilterField):
    type: Literal[FieldType.FLOAT]
    filter_value: str | None = '0.1'

    def get_value(self, value, filter_value: str | None = None):
        _filter_value = filter_value or self.filter_value
        exp = max(Decimal(_filter_value).as_tuple().exponent * -1, 1)
        adjust_value = 10 ** (-1 * (exp + 2))
        _value = Decimal(value + adjust_value).quantize(Decimal(_filter_value), rounding=ROUND_DOWN)
        _value = float(_value)
        if _value.is_integer() and re.match(r'^\d+.$', _filter_value):
            _value = str(int(_value)) + '.'
        elif _value.is_integer() and re.match(r'^\d+$', _filter_value):
            _value = str(int(_value))
        elif _value.is_integer() and re.match(r'^\d+.\d+$', _filter_value):
            _value = str(int(_value)) + '.0'
        else:
            _value = str(_value)
        return _value


class TimeField(BaseFilterField):
    type: Literal[FieldType.TIME]
    filter_value: str | None = '00:00'

    def get_value(self, value, filter_value: str | None = None):
        _filter_value = filter_value or self.filter_value
        result = ''
        if ':' not in _filter_value:
            result = value.split(':')[0]
        elif _filter_value.endswith(':'):
            result = value.split(':')[0] + ':'
        else:
            result = value
        return result


class StrField(BaseFilterField):
    type: Literal[FieldType.STR]


class BoolField(BaseFilterField):
    type: Literal[FieldType.BOOL]
    filter_value: bool = True


TypedField = Annotated[
    Annotated[BoolField, Tag(FieldType.BOOL)] | Annotated[StrField, Tag(FieldType.STR)] | Annotated[FloatField, Tag(FieldType.FLOAT)] | Annotated[TimeField, Tag(FieldType.TIME)],
    Discriminator(filter_type_discriminator)
]

ta = TypeAdapter(TypedField)


class FHBStatFilter(BaseModel):
    filter_id: PositiveInt
    filters: list[TypedField]


class Filters(RootModel):
    root: list[FHBStatFilter] | None = Field(default_factory=list)


class FHBParser(Parser):
    count_columns: int = 256
    max_time_sleep_sec: int = 1
    round_precision: str = '0.1'
    datetime_round: str = '00:00'
    count_empty_rows: int = 4
    digits_columns_start: int = 25
    enable_passability: bool
    evaluate_passability: bool
    templates: dict
    desc_dict: ClassVar[dict[str, str]] = {
        'м_2_топ': 'ТОП Лиги',
        'м_3_средн': 'Средние лиги',
        'м_4_низш': 'Низшие лиги'
    }

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._user_agent = 'Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 Mobile Safari/537.36'
        self._email = None
        self._password = None
        self._url = 'https://fhbstat.com'
        self.target_urls: defaultdict | None = defaultdict(str)
        self.file_name: str = ''
        self.from_time: str = ''
        self.to_time: str = ''
        self.user_filters: Filters | None = Filters()
        self._min_count_matches: int = 1
        self.enable_passability = False
        self.evaluate_passability = False
        self._templates = {
            '/football': ('templates.xlsx', 'Футбол исход', 0),
            '/football_24': ('templates.xlsx', 'Футбол 24', 1),
            '/hockey': ('templates.xlsx', 'Хоккей исход', 2),
            '/hockey_24': ('templates.xlsx', 'Хоккей 24', 3),
            '/football_total': ('templates.xlsx', 'Футбол тотал', 4),
            '/hockey_total': ('templates.xlsx', 'Хоккей тотал', 5),
            '/football_60': ('templates.xlsx', 'Футбол 60', 6),
        }

    @property
    def min_count_matches(self):
        return int(self._min_count_matches)

    @property
    def templates(self) -> dict:
        return self._templates

    @min_count_matches.setter
    def min_count_matches(self, value):
        if value >= 1:
            self._min_count_matches = int(value)

    @property
    def email(self):
        email = None
        if self._email:
            email = self._email
        return email

    @email.setter
    def email(self, value):
        if value:
            self._email = value

    @property
    def password(self):
        return self._password

    @password.setter
    def password(self, value):
        if value:
            self._password = value

    def stop(self):
        super().stop()
        self._email = None
        self._password = None

    def parser_log_filter(self, record):
        return __name__ == record['name']

    @property
    def columns(self):
        return list(range(1, self.count_columns))

    def clear_filters(self):
        self.user_filters.root.clear()
        self.target_urls.clear()
        self.evaluate_passability = False

    def get_filter_id(self):
        result = 1
        if self.user_filters.root:
            last_id = max(x.filter_id for x in self.user_filters.root)
            result = last_id + 1
        return result

    def add_user_filter(self, column, filter_value=None, priority=None, filter_id=None):
        exist_filter = next(
            filter(lambda x: x.filter_id == filter_id, self.user_filters.root),
            None
        )
        filter_data_dict = {
            'column': column,
            'priority': priority,
            'type': self.get_field_type(column)
        }
        if filter_value:
            filter_data_dict['filter_value'] = filter_value
        if not exist_filter:
            self.user_filters.root.append(
                FHBStatFilter(
                    filter_id=filter_id or self.get_filter_id(),
                    filters=[
                        ta.validate_python(
                            filter_data_dict
                        )
                    ]
                )
            )
        else:
            exist_column_filter = next(
                filter(lambda x: x.column == column, exist_filter.filters),
                None
            )
            if exist_column_filter:
                if filter_value:
                    exist_column_filter.filter_value = filter_value
                if priority:
                    exist_column_filter.priority = priority
            else:
                exist_filter.filters.append(
                    ta.validate_python(
                        filter_data_dict
                    )
                )

    def remove_user_filter(self, filter_id, column):
        for _filter in self.user_filters.root:
            if _filter.filter_id == filter_id:
                for _filter_ in _filter.filters:
                    if _filter_.column == column:
                        _filter.filters.remove(_filter_)
                        break

    def get_used_columns_by_filter(self, filter_id):
        result = []
        for _filter in self.user_filters.root:
            if _filter.filter_id == filter_id:
                result = [_filter_.column for _filter_ in _filter.filters]
        return result

    def download_filters(self):
        return JSONResponse(self.user_filters.model_dump())

    def upload_filters(self, json_text_filters: str):
        data = json.loads(json_text_filters)
        self.user_filters = self.user_filters.model_validate(data)

    def upload_filters_from_json(self, json_file: Path):
        self.user_filters = self.user_filters.model_validate_json(json_file.read_bytes())

    async def login(self, client: httpx.AsyncClient):
        self.status = 'Логинимся'
        cookies_file = AsyncPath('cookies.json')
        cookies = {}
        if await cookies_file.exists():
            async with cookies_file.open(encoding='utf8', mode='r') as f:
                json_content = await f.read()
                cookies = json.loads(json_content)
        response = await client.post(
            'https://fhbstat.com/авторизация',
            data={
                'posts[className]': 'вход',
                'posts[value][email]': self.email,
                'posts[value][пароль]': self.password,
                'posts[location]': 'https://fhbstat.com/авторизация',
            },
            cookies=cookies if cookies else None,
        )
        assert response.status_code == 200, 'Не удалось авторизоваться на сайте fhbstat.com'
        try:
            json_data = response.json()
        except Exception:
            self.logger.exception('Ошибка во время авторизации')
        else:
            if 'success' in json_data and 'error' in json_data['success']:
                self.status = json_data['success']['error']
                return False
            else:
                async with cookies_file.open('w', encoding='utf8') as f:
                    json_content = json.dumps(dict(client.cookies))
                    await f.write(json_content)
        return True

    async def logout(self, client: httpx.AsyncClient):
        self.status = 'Выходим'
        response = await client.post(
            'https://fhbstat.com/авторизация',
            data={
                'posts[className]': 'выход',
                'posts[value]': '',
                'posts[location]': 'https://fhbstat.com/авторизация',
            }
        )
        assert response.status_code == 200, 'Не удалось авторизоваться на сайте fhbstat.com'
        self.status = 'Вышли'

    @asynccontextmanager
    async def page_client(self, client: httpx.AsyncClient):
        try:
            is_logged = await self.login(client=client)
            if is_logged:
                yield client
            else:
                yield None
        except Exception:
            self.logger.exception('Ошибка')
        finally:
            await self.logout(client=client)

    def get_excel_template(self, path):
        return self.templates.get(path, (None, None, None))

    @classmethod
    def get_columns_by_target(cls, path):
        columns = {
            '/football': tuple(range(11, 16 + 1)),
            '/football_24': tuple(range(11, 16 + 1)),
            '/hockey': tuple(range(11, 18 + 1)),
            '/hockey_24': tuple(range(11, 18 + 1)),
            '/football_total': tuple(range(11, 16 + 1)),
            '/hockey_total': tuple(range(11, 18 + 1)),
            '/football_60': tuple(range(11, 16 + 1)),
        }
        return columns.get(path)

    def get_file_response(self, df_data, target_path) -> (FileResponse | PlainTextResponse):
        result = None
        if df_data:
            msg = f'Собрано данных: {len(df_data)}'
            self.status = msg
            df = pd.DataFrame.from_records(df_data)
            df['Дата слепка, МСК'] = self.now_msk
            columns = list(
                map(str, self.columns)
            ) + ['index', 'dt', 'Количество матчей', 'Дата слепка, МСК', 'url']
            df = df.reindex(columns=columns)
            df['Дата слепка, МСК'] = df['Дата слепка, МСК'].dt.tz_localize(None)
            older_df = pd.DataFrame(columns=columns)
            if self.file_name:
                self.path = f'files/{self.file_name}.xlsx'
                filename = f'{self.file_name}.xlsx'
            else:
                self.path = f'files/{self.name}_{self.now_msk.isoformat()}.xlsx'
                filename = f'{self.name}_{self.now_msk.isoformat()}.xlsx'
            if older_df.empty:
                full_df = df
            else:
                full_df = pd.concat((df, older_df))
            if settings.DEBUG:
                full_df.to_excel('files/debug.xlsx', index=False, columns=columns)
            full_df = full_df.reset_index(drop=True)

            template_name, sheet_name, tpl_id = self.get_excel_template(target_path)
            if all((template_name, sheet_name, tpl_id is not None)):
                fname = Path(__file__).parent.parent / Path('excel_templates') / Path(template_name)
                writer = BookWriter(fname)
                writer.jinja_env.globals.update(dir=dir, getattr=getattr)

                data = {}
                data['rows'] = df.to_dict('records')
                payload0 = {'tpl_idx': tpl_id, 'sheet_name': sheet_name, 'ctx': data}

                payloads = [payload0]
                writer.render_book2(payloads=payloads)

                workbook = writer.workbook
                sheet = workbook[sheet_name]
                start_row = None
                start_column = None
                split_column = None
                link_column = None
                for i, value in enumerate(sheet.values):
                    if all(x is None for x in value):
                        continue
                    link_name = 'ссылка'.upper()
                    if link_name in value:
                        link_column = value.index(link_name) + 1
                    count_matches_name = 'Количество матчей'
                    if '№' in value and count_matches_name in value:
                        start_row = i + 4
                        start_column = value.index('№') + 1
                        split_column = value.index(count_matches_name) + 1
                        break

                delta = 1
                if template_name == 'templates.xlsx':
                    delta = 2
                columns_by_number = list(
                    filter(
                        lambda x: sheet.cell(start_row - delta, x).value in self.get_columns_by_target(target_path),
                        range(1, link_column)
                    )
                )
                _10 = start_column
                for i in filter(lambda x: sheet.cell(start_row - delta, x).value in (10,), range(1, link_column)):
                    _10 = i

                max_rows = start_row
                for row in range(start_row + 1, sheet.max_row + 1):
                    if sheet.cell(row=row, column=start_column).value is None:
                        max_rows = row
                        break

                for col in range(start_column, _10 + 1):
                    first_row = start_row
                    end_row = first_row + len(self.user_filters.root) + 3 + self.count_empty_rows - 1
                    while end_row <= max_rows:
                        if col == start_column:
                            sheet.merge_cells(
                                start_column=col,
                                end_column=col,
                                start_row=first_row,
                                end_row=end_row
                            )
                            dimensions = list(filter(lambda x: x.hidden is False, sheet.column_dimensions.values()))
                            last_dim = dimensions[-2]
                            if dimensions[-1].min <= link_column <= dimensions[-1].max:
                                last_dim = dimensions[-2]
                            else:
                                last_dim = dimensions[-1]
                            max_column = last_dim.max
                            cell_range = CellRange(
                                min_col=col,
                                max_col=max_column,
                                min_row=first_row,
                                max_row=end_row
                            )
                            sides = ('left', 'right', 'top', 'bottom')
                            for side in sides:
                                for cell in getattr(cell_range, side, []):
                                    _cell = sheet.cell(cell[0], cell[1])
                                    other_sides = filter(lambda _side: _side != side, sides)
                                    old_border = copy(_cell.border)
                                    _border = Border(
                                        **{side: Side(border_style='thick')},
                                        **{
                                            other_side: getattr(old_border, other_side)
                                            for other_side in other_sides
                                        }
                                    )
                                    _cell.border = _border
                        else:
                            sheet.merge_cells(
                                start_column=col,
                                end_column=col,
                                start_row=first_row,
                                end_row=first_row + len(self.user_filters.root) - 1
                            )
                        first_row = end_row + 1
                        end_row = first_row + len(self.user_filters.root) + 3 + self.count_empty_rows - 1

                if link_column:
                    for row in range(start_row, max_rows + 1):
                        value = sheet.cell(row, link_column).value
                        if value and isinstance(value, str):
                            sheet.cell(row, link_column).hyperlink = value
                            sheet.cell(row, link_column).style = "Hyperlink"

                # заполнение формул
                for fn_col in range(split_column + 1, link_column):
                    for row in range(start_row, max_rows + 1):
                        if sheet.cell(row, split_column).value == '%':
                            average_columns = ','.join([
                                f'{sheet.cell(_row, fn_col).coordinate}*{sheet.cell(_row, split_column).coordinate}'
                                for _row in range(row - len(self.user_filters.root), row)
                            ])
                            sum_count_matches = (
                                f'{sheet.cell(row - len(self.user_filters.root), split_column).coordinate}:'
                                f'{sheet.cell(row - 1, split_column).coordinate}'
                            )
                            sheet.cell(row, fn_col).value = (
                                f'=ROUNDDOWN(SUM({average_columns})/SUM({sum_count_matches}),2)'
                            )
                        elif sheet.cell(row, split_column).value == 'мо':
                            sheet.cell(row, fn_col).value = (
                                f'=ROUNDDOWN(({sheet.cell(row - 2, fn_col).coordinate}/100*'
                                f'{sheet.cell(row - 1, fn_col).coordinate})-1,2)'
                            )

                for fn_col in columns_by_number:
                    for row in range(start_row, max_rows + 1):
                        if sheet.cell(row, split_column).value == '%':
                            average_columns = ','.join([
                                f'{sheet.cell(_row, fn_col).coordinate}*{sheet.cell(_row, split_column).coordinate}'
                                for _row in range(row - len(self.user_filters.root), row)
                            ])
                            sum_count_matches = (
                                f'{sheet.cell(row - len(self.user_filters.root), split_column).coordinate}:'
                                f'{sheet.cell(row - 1, split_column).coordinate}'
                            )
                            sheet.cell(row, fn_col).value = (
                                f'=ROUNDDOWN(SUM({average_columns})/SUM({sum_count_matches}),1)'
                            )

                writer.save(self.path)

                result = FileResponse(
                    self.path,
                    filename=filename
                )
            else:
                result = PlainTextResponse('Не нашли шаблон excel.')
        else:
            result = PlainTextResponse('Не собрали данных.')
        return result

    async def async_get_file_response(self, *args, **kwargs):
        return await to_thread(self.get_file_response, *args, **kwargs)

    @classmethod
    def get_head_data(cls, content):
        first_data_index = None
        names = []
        soup = BeautifulSoup(content, 'lxml')
        table_rows = list(filter(lambda tr: tr != '\n', soup.table.tbody.contents))
        first_data_row = next(
            filter(lambda tr: 'data-status' in tr.attrs, table_rows),
            None
        )
        if first_data_row:
            first_data_index = table_rows.index(first_data_row)
            names = [x.text for x in filter(lambda td: td != '\n' and td.text != '', table_rows[first_data_index - 1].contents)]
        return table_rows, first_data_index, names

    @classmethod
    def parse_head_table(cls, content):
        table_rows, first_data_index, names = cls.get_head_data(content)
        if first_data_index:
            data_rows = table_rows[3:5]
            data_list = []
            key_name = 'data-formula'
            for data in data_rows:
                data_row = {}
                for td in data.contents:
                    if td != '\n' and key_name in td.attrs:
                        key = td.attrs.get(key_name)
                        value = td.text
                        data_row[key] = float(value) if value else np.nan
                if data_row:
                    data_list.append(data_row)
            df = pd.DataFrame.from_records(data_list, columns=names + ['dt'])
            df = df.replace({None: np.nan, '': np.nan})
        else:
            df = pd.DataFrame()
        return df

    @classmethod
    def evaluate_coefficients_table(cls, df):
        head_df = pd.DataFrame(columns=df.columns).astype(np.float64)
        head_df.loc[0, '32'] = round((df['11'] > df['12']).mean() * 100, 1)
        head_df.loc[0, '33'] = round((df['11'] == df['12']).mean() * 100, 1)
        head_df.loc[0, '34'] = round((df['11'] < df['12']).mean() * 100, 1)

        _df = df.copy()
        _df = _df.drop(_df[(_df['11'] == _df['12'])].index)
        head_df.loc[0, '38'] = round((_df['11'] > _df['12']).mean() * 100, 1)
        head_df.loc[0, '39'] = round((_df['11'] < _df['12']).mean() * 100, 1)

        head_df.loc[0, '91'] = round((df['11'] + df['12'] < 1.5).mean() * 100, 1)
        head_df.loc[0, '92'] = round((df['11'] + df['12'] < 2.5).mean() * 100, 1)
        head_df.loc[0, '93'] = round((df['11'] + df['12'] < 3.5).mean() * 100, 1)

        head_df.loc[0, '94'] = round((df['11'] + df['12'] > 1.5).mean() * 100, 1)
        head_df.loc[0, '95'] = round((df['11'] + df['12'] > 2.5).mean() * 100, 1)
        head_df.loc[0, '96'] = round((df['11'] + df['12'] > 3.5).mean() * 100, 1)

        head_df.loc[0, '97'] = round((df['13'] + df['14'] < 0.5).mean() * 100, 1)
        head_df.loc[0, '98'] = round((df['13'] + df['14'] < 1.5).mean() * 100, 1)

        head_df.loc[0, '99'] = round((df['13'] + df['14'] > 0.5).mean() * 100, 1)
        head_df.loc[0, '100'] = round((df['13'] + df['14'] > 1.5).mean() * 100, 1)

        head_df.loc[0, '101'] = round((df['15'] + df['16'] < 0.5).mean() * 100, 1)
        head_df.loc[0, '102'] = round((df['15'] + df['16'] < 1.5).mean() * 100, 1)

        head_df.loc[0, '103'] = round((df['15'] + df['16'] > 0.5).mean() * 100, 1)
        head_df.loc[0, '104'] = round((df['15'] + df['16'] > 1.5).mean() * 100, 1)

        _df = df.copy()
        _df = _df.drop(_df[(_df['11'] == 1)].index)
        head_df.loc[0, '105'] = round((_df['11'] < 1).mean() * 100, 1)
        head_df.loc[0, '106'] = round((_df['11'] < 1.5).mean() * 100, 1)

        head_df.loc[0, '107'] = round((_df['11'] > 1).mean() * 100, 1)
        head_df.loc[0, '108'] = round((_df['11'] > 1.5).mean() * 100, 1)

        _df = df.copy()
        _df = _df.drop(_df[(_df['12'] == 1)].index)
        head_df.loc[0, '109'] = round((_df['12'] < 1).mean() * 100, 1)
        head_df.loc[0, '110'] = round((_df).mean() * 100, 1)

        head_df.loc[0, '111'] = round((_df['12'] > 1).mean() * 100, 1)
        head_df.loc[0, '112'] = round((_df['12'] > 1.5).mean() * 100, 1)

        head_df.loc[0, '113'] = round(
            ((df['11'] > 0.5) & (df['12'] > 0.5)).mean() * 100, 1
        )
        head_df.loc[0, '114'] = round(
            ((df['11'] < 0.5) | (df['12'] < 0.5)).mean() * 100, 1
        )

        head_df.loc[0, '115'] = round(
            ((df['13'] > 0.5) & (df['14'] > 0.5)).mean() * 100, 1
        )
        head_df.loc[0, '116'] = round(
            ((df['15'] > 0.5) & (df['16'] > 0.5)).mean() * 100, 1
        )

        head_df.loc[0, '117'] = round((df['11'] >= 0.5).mean() * 100, 1)
        head_df.loc[0, '118'] = round((df['12'] >= 0.5).mean() * 100, 1)

        head_df.loc[0, '119'] = round((df['13'] >= 0.5).mean() * 100, 1)
        head_df.loc[0, '120'] = round((df['14'] >= 0.5).mean() * 100, 1)

        head_df.loc[0, '121'] = round((df['15'] >= 0.5).mean() * 100, 1)
        head_df.loc[0, '122'] = round((df['16'] >= 0.5).mean() * 100, 1)
        return head_df

    @classmethod
    def parse_body_table(cls, content):
        table_rows, first_data_index, names = cls.get_head_data(content)
        if first_data_index:
            data_rows = table_rows[first_data_index:]
            data_list = []
            key_name = 'data-td'
            for data in data_rows:
                data_row = {}
                for td in data.contents:
                    if td != '\n' and key_name in td.attrs:
                        key = td.attrs.get(key_name)
                        value = td.text
                        if value:
                            try:
                                if value.isnumeric():
                                    data_row[key] = int(value)
                                else:
                                    data_row[key] = float(value)
                            except ValueError:
                                data_row[key] = value
                        else:
                            data_row[key] = np.nan
                _dt_str = f'{data_row.get("3")}-{data_row.get("2")}-{data_row.get("1")} {data_row.get("4")}'
                try:
                    _dt = datetime.strptime(_dt_str, '%Y-%m-%d %H:%M')
                except ValueError:
                    continue
                else:
                    data_row['dt'] = _dt
                    data_list.append(data_row)
            df = pd.DataFrame.from_records(data_list, columns=names + ['dt'])
            df = df.replace({None: np.nan, '': np.nan})
        else:
            df = pd.DataFrame()
        return df

    def get_field_type(self, value):
        if value == 4:
            return FieldType.TIME
        elif value < 11:
            return FieldType.BOOL
        elif value >= 11:
            return FieldType.FLOAT
        else:
            return FieldType.STR

    @classmethod
    def filter_df_by_time(cls, df: pd.DataFrame, from_time: str, to_time: str) -> pd.DataFrame:
        _df = df
        if not df.empty:
            if all([from_time, to_time]):
                _df = df.set_index('dt')
                _df = _df.between_time(from_time, to_time)
                _df = _df.reset_index()
            elif from_time:
                _df = df[df['dt'].dt.time >= datetime.strptime(from_time, '%H:%M').time()]
            elif to_time:
                _df = df[df['dt'].dt.time <= datetime.strptime(to_time, '%H:%M').time()]
        return _df

    @classmethod
    def get_url_params(cls, url):
        """возвращает целевой URL, параметры запроса и путь"""

        scheme, domain, path, params, query, fragment = urlparse(url)
        query_params = parse_qs(query)
        for key, value in query_params.items():
            if isinstance(value, (list, tuple)) and len(value) == 1:
                query_params[key] = value[0]
        target_url = urlunparse((scheme, domain, path, params, None, fragment))
        return target_url, query_params, path

    @classmethod
    def get_link_description(cls, url):
        _, query_params, _ = cls.get_url_params(url=url)

        for tag, value in cls.desc_dict.items():
            if tag in query_params:
                return value
        return ''

    async def _parse_page_by_filter(
        self,
        logged_client,
        browser,
        index,
        data_match,
        filters_data,
        scheme,
        domain,
        path,
        params,
        fragment,
        target_path: str
    ) -> dict:
        page_url = urlunparse((
            scheme, domain, path, params, urlencode(filters_data), fragment
        ))
        cookies = [
            {
                'name': key,
                'value': value,
                'domain': 'fhbstat.com',
                'path': '/'
            }
            for key, value in logged_client.cookies.items()
        ]
        await browser.add_cookies(cookies)
        page = await browser.new_page()
        await page.set_extra_http_headers({
            "User-Agent": self._user_agent
        })
        await page.goto(page_url)
        await page.wait_for_load_state()
        page_content = await page.content()
        df_match = self.parse_body_table(page_content)
        if not df_match.empty:
            df_match = df_match.loc[
                df_match['dt'].dt.tz_localize('Europe/Moscow') <= self.now_msk
            ]
        if self.evaluate_passability:
            head_df = self.evaluate_coefficients_table(df_match)
        else:
            head_df = self.parse_head_table(page_content)
        await page.close()
        _digits_columns_start = self.digits_columns_start if target_path != '/football_60' else 20
        columns = list(
            filter(
                lambda x: int(x) >= _digits_columns_start,
                head_df.columns[:-1]
            )
        )
        head_df_records = head_df.to_dict(orient='records')
        copy_data_match = data_match.copy()
        for h_d_r in head_df_records:
            for column_name, column_value in h_d_r.items():
                if column_name in columns:
                    copy_data_match[column_name] = column_value
        for _column in list(map(str, self.get_columns_by_target(target_path))):
            if _column in df_match.columns:
                _v = df_match[_column].mean()
                _v *= 10
                copy_data_match[_column] = (_v - _v % 1) / 10
        count_rows, _ = df_match.shape
        copy_data_match['Количество матчей'] = count_rows
        copy_data_match['index'] = index
        copy_data_match['url'] = unquote(page_url)

        return copy_data_match

    def get_last_page(self, content) -> int | None:
        last_page = None
        soup = BeautifulSoup(content, 'lxml')
        page_items = soup.find_all(lambda tag: tag.name == 'a' and 'data-pagination' in tag.attrs)
        if page_items:
            last_page = max(int(x.attrs.get('data-pagination')) for x in page_items)
        return last_page

    async def get_db(self):
        result = None
        msg = f'Открываем {self.url}'
        self.status = msg

        transport = httpx.AsyncHTTPTransport(retries=5)
        async with httpx.AsyncClient(
            follow_redirects=True,
            headers={
                'User-Agent': self._user_agent
            },
            transport=transport,
        ) as client:
            async with self.page_client(client=client) as logged_client:
                if logged_client is not None:
                    prefixes = list(filter(
                        lambda x: not Path(f'files/{x.replace("/", "")}_total_db.xlsx').exists(),
                        self.templates.keys()
                    ))
                    for path in prefixes:
                        _url = urljoin(self.url, path)
                        last_page = None
                        total_df = None
                        dfs = []
                        _target_url, query_params, target_path = self.get_url_params(_url)
                        for year in tqdm(range(2020, 2026), position=0, leave=False):
                            query_params.update({'3': year})
                            response = await logged_client.get(
                                _target_url,
                                params=query_params
                            )
                            if response.status_code == 200:
                                last_page = self.get_last_page(response.content)
                                _df = self.parse_body_table(response.content)
                                if not _df.empty:
                                    dfs.append(_df)
                            if last_page:
                                for page in tqdm(range(2, last_page + 1), position=1, leave=False):
                                    response = await logged_client.get(
                                        _target_url,
                                        params={**query_params, 'page': page}
                                    )
                                    _df = self.parse_body_table(response.content)
                                    if not _df.empty:
                                        dfs.append(_df)
                        if dfs:
                            total_df = pd.concat(dfs)
                            prefix = target_path.replace('/', '')
                            total_df.to_excel(f'files/{prefix}_total_db.xlsx', index=False)
                            result = total_df
            return result

    async def parse(self, browser):
        result = None
        msg = f'Открываем {self.url}'
        self.status = msg

        transport = httpx.AsyncHTTPTransport(retries=5)
        async with httpx.AsyncClient(
            follow_redirects=True,
            headers={
                'User-Agent': self._user_agent
            },
            transport=transport,
        ) as client, self.page_client(client=client) as logged_client:
            if logged_client is not None:
                dfs = []
                result_df_list = []
                copy_target_urls = self.target_urls.copy()
                target_path = None
                for target_url in copy_target_urls.values():
                    self.status = f'Обрабатываем ссылку {target_url}'
                    _target_url, query_params, target_path = self.get_url_params(target_url)
                    if 'page' not in query_params:
                        for page_number in count(1):
                            if page_number == 1:
                                response = await logged_client.get(
                                    _target_url,
                                    params=query_params
                                )
                            else:
                                response = await logged_client.get(
                                    _target_url,
                                    params={'page': page_number, **query_params}
                                )
                            if response.status_code == 200:
                                try:
                                    df = self.parse_body_table(response.content)
                                    df = self.filter_df_by_time(df, self.from_time, self.to_time)
                                except Exception:
                                    self.logger.exception('Ошибка сбора данных. Возможно не оплачен тариф.')
                                    self.status = 'Ошибка сбора данных. Возможно не оплачен тариф.'
                                    break
                                else:
                                    if not df.empty:
                                        dfs.append(df)
                                    else:
                                        break
                    else:
                        response = await logged_client.get(
                            _target_url,
                            params=query_params
                        )
                        if response.status_code == 200:
                            try:
                                df = self.parse_body_table(response.content)
                                df = self.filter_df_by_time(df, self.from_time, self.to_time)
                            except Exception:
                                self.logger.exception('Ошибка сбора данных. Возможно не оплачен тариф.')
                                self.status = 'Ошибка сбора данных. Возможно не оплачен тариф.'
                            else:
                                if not df.empty:
                                    dfs.append(df)
                    future_data = pd.DataFrame()
                    if dfs:
                        future_data = pd.concat(dfs)
                    data_records = future_data.to_dict(orient='records')
                    self.count_links = len(data_records)
                    for index, data_match in enumerate(self.tqdm(data_records), 1):
                        local_match_result_df = []
                        for user_filter in self.user_filters.root:
                            filters_data = {}
                            for _filter in user_filter.filters:
                                value_match = data_match.get(str(_filter.column))
                                filters_data[str(_filter.column)] = _filter.get_value(value_match)
                            for kk in self.desc_dict:
                                if kk in query_params:
                                    filters_data[kk] = '1'
                            scheme, domain, path, params, _, fragment = urlparse(_target_url)
                            priority_queues = sorted(
                                filter(
                                    lambda x: x.priority is not None,
                                    user_filter.filters
                                ),
                                key=lambda x: x.priority
                            )
                            if priority_queues:
                                _filters_data = filters_data.copy()
                                copy_data_match = {}
                                for priority_filter in priority_queues:
                                    value_match = data_match.get(str(priority_filter.column))
                                    data_exist = False
                                    for next_value in priority_filter.next_value(value_match):
                                        _filters_data[str(priority_filter.column)] = next_value
                                        copy_data_match = await self._parse_page_by_filter(
                                            logged_client,
                                            browser,
                                            index,
                                            data_match,
                                            _filters_data,
                                            scheme,
                                            domain,
                                            path,
                                            params,
                                            fragment,
                                            target_path
                                        )
                                        if copy_data_match['Количество матчей'] >= self.min_count_matches:
                                            local_match_result_df.append(copy_data_match)
                                            data_exist = True
                                            break
                                    if data_exist:
                                        break
                                if not data_exist:
                                    if copy_data_match.get('Количество матчей'):
                                        local_match_result_df.append(copy_data_match)
                                    else:
                                        local_match_result_df.append(
                                            {
                                                **{str(i): np.nan for i in self.columns},
                                                'index': index,
                                                'Количество матчей': 0,
                                                'url': unquote(
                                                    urlunparse((
                                                        scheme,
                                                        domain,
                                                        path,
                                                        params,
                                                        urlencode(filters_data),
                                                        fragment
                                                    ))
                                                ),
                                                **{str(i): data_match.get(str(i), np.nan) for i in range(11)}
                                            }
                                        )
                            else:
                                copy_data_match = await self._parse_page_by_filter(
                                    logged_client,
                                    browser,
                                    index,
                                    data_match,
                                    filters_data,
                                    scheme,
                                    domain,
                                    path,
                                    params,
                                    fragment,
                                    target_path
                                )
                                local_match_result_df.append(copy_data_match)
                        result_df_list += local_match_result_df
                        result_df_list.append({
                            **{str(i): np.nan for i in self.columns},
                            'index': index,
                            'Количество матчей': '%'
                        })
                        result_df_list.append({
                            **{
                                str(i): data_match.get(str(i))
                                for i in self.columns if i >= self.digits_columns_start
                            },
                            'index': index,
                            'Количество матчей': 'кф'
                        })
                        result_df_list.append({
                            **{str(i): np.nan for i in self.columns},
                            'index': index,
                            'Количество матчей': 'мо'
                        })
                        # Добавляем пустые строки
                        for _ in range(self.count_empty_rows):
                            result_df_list.append({
                                **{str(i): np.nan for i in self.columns},
                                'index': index
                            })
                self.status = 'Генерируем excel файл'
                result = await self.async_get_file_response(df_data=result_df_list, target_path=target_path)
                return result

    def move_name_columns(self, input_df: pd.DataFrame) -> pd.DataFrame:
        def handler_column_8(row, config: dict | None = None):
            from_column = config.get('from_column')
            result = row.get(from_column, '')
            rename_leagues = rename_leagues_country.get(row.get('7'), [])
            for rename_data in filter(lambda x: x.get('older_name', '').lower() == result.lower(), rename_leagues):
                result = rename_data.get('new_name', '')
            leagues = leagues_country.get(row.get('7'), [])
            leagues_lower = [league.lower() for league in leagues]
            if (pd.isna(result) or result.lower() not in leagues_lower) and len(leagues) >= 1:
                result = leagues[0]  # возможно будет другая логика
            return result

        russian = gettext.translation("iso3166-1", pc.LOCALES_DIR, languages=["ru"])
        russian.install()
        df = input_df.copy()
        new_columns_order = df.columns.tolist()
        leagues = {}
        fpath = Path(__file__).parent / Path('leagues.json')
        with fpath.open('r') as f:
            leagues = json.load(f)
        leagues_country = {
            league.get('label', ''): [child.get('label', '') for child in league.get('childs', [])]
            for league in leagues
            if league.get('label', '')
        }
        rename_fpath = Path(__file__).parent / Path('rename_leagues.json')
        with rename_fpath.open('r') as f:
            rename_leagues_country = json.load(f)
        columns_to_move = [
            {
                'from_column': '6',
                'to_column': '5.1',
                'words': ['Европа', 'Азия', 'Австралия', 'Южная Америка', 'Африка', 'Северная Америка'],
                'index': 'before',
                'contains': False
            },
            {
                'from_column': '7',
                'to_column': 'from_7',
                'words': [country for country in leagues_country],
                'index': 'after',
                'contains': False,
                'exclude': True,
            },
            {
                'from_column': '8',
                'to_column': '8.1',
                'func': handler_column_8,
                'index': 'after',
            }
        ]
        for data in columns_to_move:
            column = data.get('from_column')
            apply_func = data.get('func')
            if apply_func:
                df.loc[:, data['to_column']] = df.apply(partial(apply_func, config=data), axis=1)
            else:
                pattern = '|'.join(map(re.escape, data['words']))
                if data['contains']:
                    df.loc[df[column].str.contains(pattern, case=False, na=False), data['to_column']] = df[column]
                    df.loc[df[column].str.contains(pattern, case=False, na=False), column] = np.nan
                else:
                    df.loc[~df[column].str.contains(pattern, case=False, na=False), data['to_column']] = df[column]
                    df.loc[~df[column].str.contains(pattern, case=False, na=False), column] = np.nan
            if not data.get('exclude'):
                if data.get('index') == 'before':
                    insert_index = new_columns_order.index(column)
                elif data.get('index') == 'after':
                    insert_index = new_columns_order.index(column) + 1
                else:
                    insert_index = None
                if insert_index:
                    new_columns_order.insert(insert_index, data['to_column'])
        df = df.reindex(columns=new_columns_order)
        return df
